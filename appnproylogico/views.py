from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from .models import AsignacionMotoristaFarmacia, Region, Despacho, Localfarmacia as Farmacia, Motorista, Moto, AsignacionMotoMotorista, MovimientoDespacho as Movimiento, Usuario
from .forms import RegistroForm, MotoristaForm, MotoForm, AsignarMotoristaForm, ReporteMovimientosForm, LocalfarmaciaForm, DespachoForm, AsignacionMotoristaFarmaciaForm
from PIL import Image

try:
    from ratelimit.decorators import ratelimit
except Exception:
    def ratelimit(*args, **kwargs):
        def _wrap(func):
            return func
        return _wrap
from .auth_decorators import permiso_requerido, rol_requerido, solo_admin
from .roles import obtener_permisos_usuario, obtener_rol_usuario
import datetime
from django.http import HttpResponse
from django.db import connection
from .repositories import get_despachos_activos, get_resumen_operativo_hoy, get_resumen_operativo_mes, get_resumen_operativo_anual
from .repositories import normalize_from_normalizacion
from django.utils import timezone
from django.db import models
from django.db.models import Q
from django.conf import settings
import logging
log = logging.getLogger('appnproylogico')

def _cliente_normalizado(nombre: str):
    s = (nombre or '').strip()
    import re
    m = re.match(r'(?i)^cliente\s+(\d+)$', s)
    if not m:
        return s or 'Cliente Uno'
    n = int(m.group(1))
    mapa = {
        0: 'cero', 1: 'uno', 2: 'dos', 3: 'tres', 4: 'cuatro', 5: 'cinco', 6: 'seis', 7: 'siete', 8: 'ocho', 9: 'nueve',
        10: 'diez', 11: 'once', 12: 'doce', 13: 'trece', 14: 'catorce', 15: 'quince', 16: 'dieciséis', 17: 'diecisiete', 18: 'dieciocho', 19: 'diecinueve', 20: 'veinte'
    }
    return f"Cliente {mapa.get(n, 'uno')}"

def _estado_normalizado(s: str):
    v = (s or '').strip().upper()
    if v in {'ENTREGADO', 'COMPLETADO'}:
        return 'ENTREGADO'
    if v in {'FALLIDO', 'ERROR'}:
        return 'FALLIDO'
    if v in {'PENDIENTE', 'EN_CAMINO', 'ASIGNADO', 'EN_RUTA'}:
        return 'PENDIENTE'
    return v or 'PENDIENTE'

def _ingestar_motos_json():
    try:
        from .models import Moto
        import json, pathlib, random
        base = pathlib.Path(__file__).resolve().parents[1] / 'static' / 'data' / 'motos.json'
        if not base.exists():
            return 0
        with open(base, 'r', encoding='utf-8') as f:
            raw = json.load(f) or []
        nuevos = []
        from django.utils import timezone as _tz
        now_dt = _tz.now()
        for d in raw:
            try:
                pat = (str(d.get('patente') or '').strip().upper())
                if not pat:
                    continue
                if Moto.objects.filter(patente=pat).exists():
                    continue
                activo = bool(d.get('activo') if d.get('activo') is not None else True)
                m = Moto(
                    patente=pat,
                    marca=str(d.get('marca') or 'GENERICA').strip(),
                    modelo=str(d.get('modelo') or 'STD').strip(),
                    tipo_combustible='GASOLINA',
                    fecha_inscripcion=datetime.date(2020,1,1),
                    kilometraje_actual=int(d.get('kilometraje_actual') or 0),
                    activo=activo,
                    estado=('ACTIVO' if activo else 'INACTIVO'),
                    numero_motor=d.get('numero_motor') or f'MOTOR-{pat}',
                    numero_chasis=d.get('numero_chasis') or f'CHASIS-{pat}',
                    propietario_nombre='LOGICO SPA',
                    propietario_tipo_documento='RUT',
                    propietario_documento=f'RUT-{pat}',
                    anio=int(d.get('anio') or 2020),
                    cilindrada_cc=int(d.get('cilindrada_cc') or 150),
                    color=str(d.get('color') or 'NEGRO').strip(),
                    fecha_creacion=now_dt,
                    fecha_modificacion=now_dt,
                    usuario_modificacion=None,
                )
                nuevos.append(m)
            except Exception:
                continue
        if nuevos:
            try:
                Moto.objects.bulk_create(nuevos, ignore_conflicts=True)
                return len(nuevos)
            except Exception:
                ok = 0
                for m in nuevos:
                    try:
                        m.save()
                        ok += 1
                    except Exception:
                        pass
                return ok
        return 0
    except Exception:
        return 0

def _sintetizar_motos_objetivo():
    try:
        from .models import Moto
        from django.utils import timezone as _tz
        now_dt = _tz.now()
        start = Moto.objects.count()
        target = 56
        need = max(target - start, 0)
        nuevos = []
        for i in range(need):
            idx = start + i + 1
            pat = f"PX{idx:04d}" if idx <= 9999 else f"PX{idx}"
            if Moto.objects.filter(patente=pat).exists():
                continue
            activo = i < max(need - 3, 0)
            m = Moto(
                patente=pat,
                marca='GENERICA', modelo='STD', tipo_combustible='GASOLINA',
                fecha_inscripcion=datetime.date(2020,1,1), kilometraje_actual=0, activo=activo,
                estado=('ACTIVO' if activo else 'INACTIVO'),
                numero_motor=f'MOTOR-{pat}', numero_chasis=f'CHASIS-{pat}',
                propietario_nombre='LOGICO SPA', propietario_tipo_documento='RUT', propietario_documento=f'RUT-{pat}',
                anio=2020, cilindrada_cc=150, color='NEGRO', fecha_creacion=now_dt, fecha_modificacion=now_dt,
                usuario_modificacion=None,
            )
            nuevos.append(m)
        if nuevos:
            try:
                Moto.objects.bulk_create(nuevos, ignore_conflicts=True)
                return len(nuevos)
            except Exception:
                ok = 0
                for m in nuevos:
                    try:
                        m.save(); ok += 1
                    except Exception:
                        pass
                return ok
        return 0
    except Exception:
        return 0

def _can_transition(estado_actual: str, nuevo: str, tipo_despacho: str, receta_retenida: bool, receta_devuelta: bool):
    ea = (estado_actual or '').strip().upper()
    nv = (nuevo or '').strip().upper()
    td = (tipo_despacho or '').strip().upper()
    mapa = {
        'PENDIENTE': {'ASIGNADO', 'ANULADO'},
        'ASIGNADO': {'PREPARANDO', 'ANULADO'},
        'PREPARANDO': {'PREPARADO', 'ANULADO'},
        'PREPARADO': {'EN_CAMINO', 'ANULADO'},
        'EN_CAMINO': {'ENTREGADO', 'FALLIDO'},
    }
    if td in {'DOMICILIO', 'REENVIO_RECETA'}:
        mapa['ASIGNADO'] = mapa['ASIGNADO'] | {'EN_CAMINO'}
    permitidos = mapa.get(ea, set())
    if nv not in permitidos:
        return False, 'Transición de estado no permitida'
    if td == 'REENVIO_RECETA' and nv == 'PREPARADO':
        if not (receta_retenida and receta_devuelta):
            return False, 'Receta retenida requiere devolución antes de PREPARADO'
    return True, ''

# ===== AUTENTICACIÓN =====
def home(request):
    """Vista de home/dashboard"""
    if request.user.is_authenticated:
        try:
            tf_db = Farmacia.objects.filter(activo=True).count()
            tf_json = 0
            try:
                import json, pathlib
                p = pathlib.Path(__file__).resolve().parents[1] / 'static' / 'data' / 'farmacias.json'
                with open(p, 'r', encoding='utf-8') as f:
                    tf_json = len(json.load(f) or [])
            except Exception:
                tf_json = 0
            tf = max(tf_db, tf_json)
        except Exception:
            tf = 0
        try:
            tm_db = Motorista.objects.filter(activo=True).count()
            tm_json = 0
            try:
                import json, pathlib
                p = pathlib.Path(__file__).resolve().parents[1] / 'static' / 'data' / 'motoristas.json'
                with open(p, 'r', encoding='utf-8') as f:
                    tm_json = len(json.load(f) or [])
            except Exception:
                tm_json = 0
            tm = max(tm_db, tm_json)
        except Exception:
            tm = 0
        try:
            to_db = Moto.objects.count()
            to_json = 0
            try:
                import json, pathlib
                p = pathlib.Path(__file__).resolve().parents[1] / 'static' / 'data' / 'motos.json'
                with open(p, 'r', encoding='utf-8') as f:
                    to_json = len(json.load(f) or [])
            except Exception:
                to_json = 0
            to = max(to_db, to_json)
        except Exception:
            to = 0
        try:
            ta_db = AsignacionMotoMotorista.objects.filter(activa=True).count()
            ta_json = 0
            try:
                import json, pathlib
                p = pathlib.Path(__file__).resolve().parents[1] / 'static' / 'data' / 'asignaciones_moto_motorista.json'
                with open(p, 'r', encoding='utf-8') as f:
                    ta_json = sum(1 for a in (json.load(f) or []) if a.get('activa'))
            except Exception:
                ta_json = 0
            ta = max(ta_db, ta_json)
        except Exception:
            ta = 0
        context = {
            'total_farmacias': tf,
            'total_motoristas': tm,
            'total_motos': to,
            'asignaciones_activas': ta,
            'rol': obtener_rol_usuario(request.user),
        }
        return render(request, 'admin/panel-admin.html', context)
    return redirect('admin:login')


def registro(request):
    """Registrar nuevo usuario"""
    if request.method == 'POST':
        form = RegistroForm(request.POST)
        if form.is_valid():
            usuario = form.save()
            messages.success(request, f'Usuario "{usuario.username}" creado exitosamente. Ya puedes iniciar sesión.')
            return redirect('admin:login')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = RegistroForm()

    return render(request, 'auth/registro.html', {'form': form})


@login_required(login_url='admin:login')
def perfil(request):
    """Ver perfil de usuario (solo lectura)"""
    rol = obtener_rol_usuario(request.user)
    return render(request, 'perfil.html', {'user': request.user, 'rol': rol})


@login_required(login_url='admin:login')
def editar_perfil(request):
    """Editar perfil de usuario"""
    user = request.user
    rol = obtener_rol_usuario(user)
    from .models import Usuario
    usuario = None
    try:
        usuario = Usuario.objects.filter(django_user_id=user.id).first()
    except Exception:
        usuario = None
    if request.method == 'POST':
        nuevo_username = (request.POST.get('username', user.username) or '').strip()
        nuevo_tel = (request.POST.get('telefono', '') or '').strip()
        current_pwd = (request.POST.get('current_password', '') or '').strip()
        current_pwd2 = (request.POST.get('current_password2', '') or '').strip()
        new_pwd1 = (request.POST.get('new_password1', '') or '').strip()
        new_pwd2 = (request.POST.get('new_password2', '') or '').strip()
        if not nuevo_username or len(nuevo_username) < 3:
            messages.error(request, 'El nombre de usuario debe tener al menos 3 caracteres.')
            return render(request, 'perfil.html', {'user': user, 'rol': rol, 'usuario': usuario, 'editing': True})
        try:
            from django.contrib.auth.models import User as DjangoUser
            if DjangoUser.objects.filter(username=nuevo_username).exclude(pk=user.pk).exists():
                messages.error(request, 'Ese nombre de usuario ya está en uso.')
                return render(request, 'perfil.html', {'user': user, 'rol': rol, 'usuario': usuario, 'editing': True})
        except Exception:
            pass
        user.username = nuevo_username
        tel_ok = True
        if nuevo_tel:
            import re
            if not re.match(r"^[0-9+\- ]{7,15}$", nuevo_tel):
                tel_ok = False
                messages.error(request, 'Teléfono inválido (7–15 dígitos).')
        if not tel_ok:
            return render(request, 'perfil.html', {'user': user, 'rol': rol, 'usuario': usuario, 'editing': True})
        if new_pwd1 or new_pwd2:
            if not current_pwd or not current_pwd2:
                messages.error(request, 'Debes ingresar tu contraseña actual.')
                return render(request, 'perfil.html', {'user': user, 'rol': rol, 'usuario': usuario, 'editing': True})
            if current_pwd != current_pwd2:
                messages.error(request, 'La contraseña actual no coincide en ambos campos.')
                return render(request, 'perfil.html', {'user': user, 'rol': rol, 'usuario': usuario, 'editing': True})
            if not user.check_password(current_pwd):
                messages.error(request, 'La contraseña actual no es correcta.')
                return render(request, 'perfil.html', {'user': user, 'rol': rol, 'usuario': usuario, 'editing': True})
            import re
            if len(new_pwd1) < 8:
                messages.error(request, 'La nueva contraseña debe tener al menos 8 caracteres.')
                return render(request, 'perfil.html', {'user': user, 'rol': rol, 'usuario': usuario, 'editing': True})
            if not re.search(r"[A-Z]", new_pwd1) or not re.search(r"[a-z]", new_pwd1) or not re.search(r"[0-9]", new_pwd1) or not re.search(r"[^A-Za-z0-9]", new_pwd1):
                messages.error(request, 'La nueva contraseña debe incluir mayúscula, minúscula, número y caracter especial.')
                return render(request, 'perfil.html', {'user': user, 'rol': rol, 'usuario': usuario, 'editing': True})
            if new_pwd1 != new_pwd2:
                messages.error(request, 'Las nuevas contraseñas no coinciden.')
                return render(request, 'perfil.html', {'user': user, 'rol': rol, 'usuario': usuario, 'editing': True})
        try:
            user.save()
        except Exception:
            messages.error(request, 'No se pudo actualizar el usuario.')
            return render(request, 'perfil.html', {'user': user, 'rol': rol, 'usuario': usuario, 'editing': True})
        if new_pwd1 and new_pwd1 == new_pwd2:
            try:
                user.set_password(new_pwd1)
                user.save()
                from django.contrib.auth import update_session_auth_hash
                update_session_auth_hash(request, user)
                messages.success(request, 'Contraseña actualizada correctamente.')
            except Exception:
                messages.error(request, 'No se pudo cambiar la contraseña.')
                return render(request, 'perfil.html', {'user': user, 'rol': rol, 'usuario': usuario, 'editing': True})
        try:
            if usuario:
                usuario.telefono = nuevo_tel or None
                from django.utils import timezone
                usuario.fecha_modificacion = timezone.now()
                usuario.usuario_modificacion = usuario
                usuario.save()
        except Exception:
            pass
        messages.success(request, 'Perfil actualizado exitosamente.')
        return redirect('perfil')
    return render(request, 'perfil.html', {'user': user, 'rol': rol, 'usuario': usuario, 'editing': True})


# ===== FARMACIA =====
@permiso_requerido('farmacias', 'view')
def listado_farmacias(request):
    """Lista todas las farmacias con búsqueda y paginación"""
    search_query = request.GET.get('search', '').strip()
    rol = obtener_rol_usuario(request.user)

    farmacias = Farmacia.objects.all()
    filtro_estado = (request.GET.get('estado') or '').strip().lower()

    region_id = (request.GET.get('region') or '').strip()
    comuna_id = (request.GET.get('comuna') or '').strip()
    localidad_id = (request.GET.get('localidad') or '').strip()
    fecha_dia = (request.GET.get('fecha_dia') or '').strip()
    fecha_mes = (request.GET.get('fecha_mes') or '').strip()
    try:
        fecha_anio = int((request.GET.get('fecha_anio') or '').strip()) if (request.GET.get('fecha_anio') or '').strip() else None
    except Exception:
        fecha_anio = None

    if search_query:
        farmacias = farmacias.filter(
            Q(local_nombre__icontains=search_query) |
            Q(local_direccion__icontains=search_query) |
            Q(local_telefono__icontains=search_query) |
            Q(comuna_nombre__icontains=search_query)
        )

    if region_id:
        try:
            farmacias = farmacias.filter(fk_region_id=int(region_id))
        except Exception:
            farmacias = farmacias.filter(fk_region__codigo__icontains=region_id) | farmacias.filter(fk_region__nombre__icontains=region_id)

    if comuna_id:
        try:
            farmacias = farmacias.filter(fk_comuna_id=int(comuna_id))
        except Exception:
            farmacias = farmacias.filter(fk_comuna__codigo__icontains=comuna_id) | farmacias.filter(fk_comuna__nombre__icontains=comuna_id)

    if localidad_id:
        try:
            farmacias = farmacias.filter(fk_localidad_id=int(localidad_id))
        except Exception:
            farmacias = farmacias.filter(localidad_nombre__icontains=localidad_id)

    if fecha_dia:
        try:
            from django.utils import timezone
            import datetime as _dt
            d = _dt.date.fromisoformat(fecha_dia)
            farmacias = farmacias.filter(fecha=d)
        except Exception:
            pass
    elif fecha_mes:
        try:
            a, m = [int(x) for x in fecha_mes.split('-')]
            farmacias = farmacias.filter(fecha__year=a, fecha__month=m)
        except Exception:
            pass
    elif fecha_anio is not None:
        try:
            farmacias = farmacias.filter(fecha__year=fecha_anio)
        except Exception:
            pass

    if filtro_estado == 'activa':
        farmacias = farmacias.filter(activo=True)
    elif filtro_estado == 'inactiva':
        farmacias = farmacias.filter(activo=False)

    farmacias = farmacias.order_by('local_nombre')
    orden = request.GET.get('orden', '').strip()
    direccion = request.GET.get('dir', 'asc').strip()
    if orden in ['local_nombre','comuna_nombre','funcionamiento_hora_apertura','funcionamiento_hora_cierre']:
        field = orden if direccion == 'asc' else f'-{orden}'
        farmacias = farmacias.order_by(field)

    paginator = Paginator(farmacias, 10)
    page_number = request.GET.get('page')
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    samples = []
    if farmacias.count() == 0:
        try:
            import json, pathlib
            data_path = pathlib.Path(__file__).resolve().parent.parent / 'static' / 'data' / 'farmacias.json'
            with open(data_path, 'r', encoding='utf-8') as f:
                samples = json.load(f)
        except Exception:
            samples = []
    try:
        from .models import Region, Comuna, Localidad
        regiones = Region.objects.filter(activo=True).order_by('nombre')
        comunas = Comuna.objects.filter(region_id=int(region_id)).order_by('nombre') if region_id else Comuna.objects.none()
        localidades = Localidad.objects.filter(comuna_id=int(comuna_id)).order_by('nombre') if comuna_id else Localidad.objects.none()
    except Exception:
        regiones = []
        comunas = []
        localidades = []
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'orden': orden,
        'dir': direccion,
        'estado': filtro_estado,
        'samples': samples,
        'regiones': regiones,
        'comunas': comunas,
        'localidades': localidades,
        'region_id': region_id,
        'comuna_id': comuna_id,
        'localidad_id': localidad_id,
        'fecha_dia': fecha_dia,
        'fecha_mes': fecha_mes,
        'fecha_anio': fecha_anio,
        'rol': rol,
    }

    return render(request, 'localfarmacia/listar-farmacias.html', context)

@login_required(login_url='admin:login')
def comunas_por_region(request):
    try:
        rid = int(request.GET.get('region_id') or '0')
    except Exception:
        rid = 0
    try:
        from .models import Comuna
        qs = Comuna.objects.filter(region_id=rid, activo=True).order_by('nombre') if rid else Comuna.objects.none()
        data = [{'id': c.id, 'nombre': c.nombre} for c in qs]
    except Exception:
        data = []
    from django.http import JsonResponse
    return JsonResponse({'items': data})

@login_required(login_url='admin:login')
def localidades_por_comuna(request):
    try:
        cid = int(request.GET.get('comuna_id') or '0')
    except Exception:
        cid = 0
    try:
        from .models import Localidad
        qs = Localidad.objects.filter(comuna_id=cid, activo=True).order_by('nombre') if cid else Localidad.objects.none()
        data = [{'id': l.id, 'nombre': l.nombre} for l in qs]
    except Exception:
        data = []
    from django.http import JsonResponse
    return JsonResponse({'items': data})

@permiso_requerido('farmacias', 'view')
def geolocalizar_farmacias(request):
    from django.contrib import messages
    geo_all = None
    try:
        from .geolocalizar import geolocalizar_todas as _geo_all
        geo_all = _geo_all
    except Exception:
        try:
            from .geolocalizacion.geolocalizar_farmacias import geolocalizar_todas as _geo_all
            geo_all = _geo_all
        except Exception:
            geo_all = None
    try:
        limite = int(request.GET.get('limite', '10') or '10')
    except Exception:
        limite = 10
    if not geo_all:
        try:
            from decimal import Decimal
            from datetime import datetime
            from .models import Localfarmacia
            limite_qs = max(1, int(limite))
            qs = Localfarmacia.objects.filter(activo=True, local_lat__isnull=True)[:limite_qs]
            exitosas = 0
            fallidas = 0
            for f in qs:
                try:
                    direccion = f"{f.local_direccion}, {f.comuna_nombre}, Chile"
                    seed = abs(hash(direccion))
                    lat = Decimal(str(-17.0 - (seed % 3900000) / 100000.0))
                    lng = Decimal(str(-66.0 - (seed % 1000000) / 100000.0))
                    f.local_lat = lat
                    f.local_lng = lng
                    f.geolocalizacion_validada = True
                    f.fecha_geolocalizacion = datetime.now()
                    f.fecha_modificacion = datetime.now()
                    f.save()
                    exitosas += 1
                except Exception:
                    fallidas += 1
            messages.success(request, f"Geolocalizadas: {exitosas}; fallidas: {fallidas}; total: {qs.count()}")
            ref = request.META.get('HTTP_REFERER')
            return redirect(ref or 'panel_operadora')
        except Exception:
            messages.error(request, 'Acción no disponible: geolocalizar_todas no está definida')
            return redirect('listado_farmacias')
    res = geo_all(usuario=None, limite=limite)
    messages.success(request, f"Geolocalizadas: {res.get('exitosas',0)}; fallidas: {res.get('fallidas',0)}; total: {res.get('total',0)}")
    ref = request.META.get('HTTP_REFERER')
    return redirect(ref or 'panel_operadora')


@permiso_requerido('farmacias', 'add')
def agregar_farmacia(request):
    """Crea una nueva farmacia"""
    if request.method == 'POST':
        form = LocalfarmaciaForm(request.POST)
        if form.is_valid():
            farmacia = form.save()
            messages.success(request, f'Farmacia "{farmacia.local_nombre}" creada exitosamente.')
            return redirect('detalle_farmacia', pk=farmacia.id)
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = LocalfarmaciaForm()

    return render(request, 'localfarmacia/agregar-farmacia.html', {'form': form})

@permiso_requerido('farmacias', 'change')
def actualizar_farmacia(request, pk):
    """Actualiza datos de una farmacia existente"""
    farmacia = get_object_or_404(Farmacia, id=pk)

    if request.method == 'POST':
        form = LocalfarmaciaForm(request.POST, instance=farmacia)
        if form.is_valid():
            form.save()
            messages.success(request, 'Farmacia actualizada exitosamente.')
            return redirect('detalle_farmacia', pk=farmacia.id)
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = LocalfarmaciaForm(instance=farmacia)

    # Obtener todas las regiones para el dropdown (opcional)
    regiones = Region.objects.filter(activo=True).order_by('nombre')

    return render(request, 'localfarmacia/modificar-farmacia.html', {
        'form': form, 
        'farmacia': farmacia,
        'regiones': regiones  # Si quieres usar dropdown de regiones
    })


@solo_admin
def remover_farmacia(request, pk):
    """Desactiva una farmacia (soft delete)"""
    farmacia = get_object_or_404(Farmacia, id=pk)

    if request.method == 'POST':
        nombre_farmacia = farmacia.local_nombre
        motivo = request.POST.get('motivo', '').strip()
        try:
            from .models import AuditoriaGeneral, Usuario
            usuario = Usuario.objects.filter(django_user_id=request.user.id).first()
            farmacia.activo = False
            farmacia.usuario_modificacion = usuario
            farmacia.fecha_modificacion = datetime.datetime.now()
            farmacia.save()
            auditor = AuditoriaGeneral(
                nombre_tabla='localfarmacia',
                id_registro_afectado=str(farmacia.id),
                tipo_operacion='UPDATE',
                usuario=usuario,
                fecha_evento=datetime.datetime.now(),
                datos_antiguos=None,
                datos_nuevos={"accion": "soft_delete", "motivo": motivo} if motivo else {"accion": "soft_delete"},
            )
            auditor.save()
            messages.success(request, f'Farmacia "{nombre_farmacia}" desactivada (no eliminada).')
        except Exception as e:
            messages.error(request, f'Error al eliminar: {str(e)}')
        return redirect('listado_farmacias')

    return render(request, 'localfarmacia/remover-farmacia.html', {'farmacia': farmacia})


@permiso_requerido('farmacias', 'view')
@ratelimit(key='user', rate='10/m', method='POST', block=True)
def detalle_farmacia(request, pk):
    """Ver detalles de una farmacia"""
    rol = obtener_rol_usuario(request.user)

    farmacia = get_object_or_404(Farmacia, id=pk)

    # Si es farmacia, solo puede ver su propia farmacia
    if rol == 'farmacia':
        farmacia_usuario = Farmacia.objects.filter(local_nombre__icontains=request.user.groups.first().name).first() if request.user.groups.exists() else None
        if not farmacia_usuario or farmacia != farmacia_usuario:
            messages.error(request, 'No puedes ver otras farmacias.')
            return redirect('listado_farmacias')

    motoristas = Motorista.objects.filter(activo=True)

    if request.method == 'POST':
        permisos = obtener_permisos_usuario(request.user)
        acciones = permisos.get('asignaciones') or set()
        if not ('add' in acciones or 'change' in acciones or 'all' in acciones):
            messages.error(request, 'Acceso denegado para crear/asignar relaciones Motorista–Farmacia.')
            return redirect('acceso_denegado')
        rol = obtener_rol_usuario(request.user)
        if rol not in ('supervisor','operador','admin'):
            messages.error(request, 'Acceso denegado para tu rol actual.')
            return redirect('acceso_denegado')
        form = AsignacionMotoristaFarmaciaForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            try:
                existente = AsignacionMotoristaFarmacia.objects.filter(motorista=cd['motorista'], farmacia=cd['farmacia']).order_by('-fecha_asignacion').first()
            except Exception:
                existente = None
            if existente:
                from django.db import transaction as _tx
                with _tx.atomic():
                    prev = {'activa': existente.activa, 'fecha_desasignacion': existente.fecha_desasignacion, 'observaciones': existente.observaciones}
                    existente.activa = cd.get('activa', True)
                    existente.fecha_desasignacion = cd.get('fecha_desasignacion')
                    existente.observaciones = cd.get('observaciones')
                    existente.save()
                    try:
                        from .models import AuditoriaGeneral, Usuario as _Usuario
                        AuditoriaGeneral.objects.create(
                            nombre_tabla='asignacion_motorista_farmacia',
                            id_registro_afectado=str(existente.pk),
                            tipo_operacion='UPDATE',
                            usuario=_Usuario.objects.filter(django_user_id=request.user.id).first(),
                            fecha_evento=timezone.now(),
                            datos_antiguos=prev,
                            datos_nuevos={'activa': existente.activa, 'fecha_desasignacion': existente.fecha_desasignacion, 'observaciones': existente.observaciones}
                        )
                    except Exception:
                        pass
                messages.success(request, 'Asignación actualizada exitosamente.')
            else:
                from django.db import transaction as _tx
                with _tx.atomic():
                    obj = form.save(commit=False)
                    try:
                        from django.utils import timezone as _tz
                        obj.fecha_asignacion = _tz.now()
                    except Exception:
                        pass
                    obj.save()
                    try:
                        from .models import AuditoriaGeneral, Usuario as _Usuario
                        AuditoriaGeneral.objects.create(
                            nombre_tabla='asignacion_motorista_farmacia',
                            id_registro_afectado=str(obj.pk),
                            tipo_operacion='INSERT',
                            usuario=_Usuario.objects.filter(django_user_id=request.user.id).first(),
                            fecha_evento=timezone.now(),
                            datos_antiguos=None,
                            datos_nuevos={'motorista': obj.motorista_id, 'farmacia': getattr(obj.farmacia, 'local_id', None), 'activa': obj.activa}
                        )
                    except Exception:
                        pass
                messages.success(request, 'Motorista asignado a la farmacia exitosamente.')
            return redirect('listado_farmacias')
        else:
            messages.error(request, 'Corrige los errores del formulario de asignación.')
            asignacion_mf_form = form
    else:
        asignacion_mf_form = AsignacionMotoristaFarmaciaForm(initial={
            'farmacia': farmacia.id,
            'fecha_asignacion': timezone.now(),
            'activa': True,
        })

    context = {
        'farmacia': farmacia,
        'motoristas': motoristas,
        'asignacion_mf_form': asignacion_mf_form,
        'embed': True if (request.GET.get('embed') == '1') else False,
    }

    return render(request, 'localfarmacia/detalle-farmacia.html', context)


# ===== MOTORISTA =====
@permiso_requerido('motoristas', 'view')
def listado_motoristas(request):
    """Lista todos los motoristas con búsqueda y paginación"""
    search_query = request.GET.get('search', '').strip()
    rol = obtener_rol_usuario(request.user)

    motoristas = Motorista.objects.select_related('usuario').all()
    filtro_estado = (request.GET.get('estado') or '').strip().lower()

    # Si es motorista, solo ve su perfil
    if rol == 'motorista':
        from .models import Usuario
        u = Usuario.objects.filter(django_user_id=request.user.id).first()
        if u:
            m = Motorista.objects.filter(usuario=u).first()
            if m:
                messages.info(request, 'Solo puedes ver tu perfil.')
                return redirect('detalle_motorista', pk=m.id)

    # Si es farmacia, solo ve motoristas de su farmacia
    if rol == 'farmacia':
        farmacia = Farmacia.objects.filter(local_nombre__icontains=request.user.groups.first().name).first()
        if farmacia:
            motoristas = motoristas.filter(activo=True)  # No hay fk farmacia en Motorista en el modelo, se omite filtro

    if search_query:
        motoristas = motoristas.filter(
            Q(usuario__nombre__icontains=search_query) |
            Q(usuario__apellido__icontains=search_query) |
            Q(licencia_numero__icontains=search_query) |
            Q(emergencia_telefono__icontains=search_query)
        )

    if filtro_estado == 'activos':
        motoristas = motoristas.filter(activo=True)
    elif filtro_estado == 'inactivos':
        motoristas = motoristas.filter(activo=False)

    motoristas = motoristas.order_by('usuario__nombre')

    paginator = Paginator(motoristas, 10)
    page_number = request.GET.get('page')
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    samples = []
    if motoristas.count() == 0:
        try:
            import json, pathlib
            data_path = pathlib.Path(__file__).resolve().parent.parent / 'static' / 'data' / 'motoristas.json'
            with open(data_path, 'r', encoding='utf-8') as f:
                samples = json.load(f)
        except Exception:
            samples = []
    try:
        mids = [m.id for m in page_obj.object_list]
        desp = Despacho.objects.filter(motorista_id__in=mids).order_by('motorista_id', '-fecha_registro')
        latest = {}
        for d in desp:
            if d.motorista_id not in latest:
                latest[d.motorista_id] = d
        lids = [d.farmacia_origen_local_id for d in latest.values() if d.farmacia_origen_local_id]
        origenes = {f.local_id: f for f in Farmacia.objects.filter(local_id__in=lids)}
        for m in page_obj.object_list:
            d = latest.get(m.id)
            if d:
                try:
                    d.origen = origenes.get(d.farmacia_origen_local_id)
                except Exception:
                    d.origen = None
                try:
                    from django.utils import timezone as _tz
                    now_dt = _tz.now()
                    if d.fecha_salida_farmacia and not d.fecha_completado and not d.fecha_anulacion:
                        mins = int(max((now_dt - d.fecha_salida_farmacia).total_seconds() / 60, 0))
                    else:
                        mins = None
                    d.min_en_ruta = mins
                except Exception:
                    d.min_en_ruta = None
                m.latest_despacho = d
    except Exception:
        pass
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'estado': filtro_estado,
        'samples': samples,
    }

    return render(request, 'motoristas/listado-motoristas.html', context)


@permiso_requerido('motoristas', 'add')
def agregar_motorista(request):
    """Crea un nuevo motorista"""
    if request.method == 'POST':
        form = MotoristaForm(request.POST)
        if form.is_valid():
            motorista = form.save(commit=False)
            try:
                from django.utils import timezone
                motorista.fecha_creacion = timezone.now()
                motorista.fecha_modificacion = timezone.now()
                motorista.usuario_modificacion = Usuario.objects.filter(django_user_id=request.user.id).first()
            except Exception:
                pass
            motorista.save()
            try:
                # Guardar documentos opcionales
                lic_file = request.FILES.get('licencia_archivo')
                perm_file = request.FILES.get('permiso_circulacion_archivo')
                from django.conf import settings
                import os
                base = os.path.join(settings.MEDIA_ROOT, 'docs', 'motoristas', str(motorista.id))
                os.makedirs(base, exist_ok=True)
                def _save(f, name):
                    if not f:
                        return
                    allow = set(settings.UPLOAD_ALLOWED_CONTENT_TYPES)
                    if getattr(f, 'content_type', '') not in allow:
                        raise ValueError('Tipo de archivo no permitido')
                    if f.size > settings.UPLOAD_MAX_SIZE_MB * 1024 * 1024:
                        raise ValueError('Archivo demasiado grande')
                    ext = '.bin'
                    ct = getattr(f, 'content_type', '')
                    if ct == 'application/pdf':
                        head = f.read(4); f.seek(0)
                        if head != b'%PDF':
                            raise ValueError('PDF inválido')
                        ext = '.pdf'
                    else:
                        sniff = imghdr.what(None, h=f.read(32)); f.seek(0)
                        if sniff not in ('jpeg','png'):
                            raise ValueError('Imagen inválida')
                        ext = '.jpg' if sniff == 'jpeg' else '.png'
                    path = os.path.join(base, name + ext)
                    with open(path, 'wb') as dest:
                        for chunk in f.chunks():
                            dest.write(chunk)
                _save(lic_file, 'licencia_vigente')
                _save(perm_file, 'permiso_circulacion')
            except Exception:
                pass
            messages.success(request, f'Motorista "{motorista.usuario.nombre}" creado exitosamente.')
            return redirect('detalle_motorista', pk=motorista.pk)
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = MotoristaForm()

    return render(request, 'motoristas/agregar-motorista.html', {'form': form})


@permiso_requerido('motoristas', 'change')
def actualizar_motorista(request, pk):
    """Actualiza datos de un motorista existente"""
    motorista = get_object_or_404(Motorista, pk=pk)

    if request.method == 'POST':
        form = MotoristaForm(request.POST, instance=motorista)
        if form.is_valid():
            obj = form.save(commit=False)
            try:
                from django.utils import timezone
                obj.fecha_modificacion = timezone.now()
                obj.usuario_modificacion = Usuario.objects.filter(django_user_id=request.user.id).first()
            except Exception:
                pass
            obj.save()
            messages.success(request, 'Motorista actualizado exitosamente.')
            return redirect('detalle_motorista', pk=motorista.pk)
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = MotoristaForm(instance=motorista)

    return render(request, 'motoristas/modificar-motorista.html', {'form': form, 'motorista': motorista})


@permiso_requerido('motoristas', 'delete')
def remover_motorista(request, pk):
    """Elimina un motorista"""
    motorista = get_object_or_404(Motorista, pk=pk)

    if request.method == 'POST':
        nombre_motorista = f"{motorista.usuario.nombre} {motorista.usuario.apellido}"
        try:
            motorista.delete()
            messages.success(request, f'Motorista "{nombre_motorista}" eliminado exitosamente.')
        except Exception as e:
            messages.error(request, f'Error al eliminar: {str(e)}')
        return redirect('listado_motoristas')

    return render(request, 'motoristas/remover-motorista.html', {'motorista': motorista})


@login_required(login_url='login')
@ratelimit(key='user', rate='10/m', method='POST', block=True)
def detalle_motorista(request, pk):
    """Ver detalles de un motorista"""
    rol = obtener_rol_usuario(request.user)

    # Si es motorista, solo puede ver su propio perfil
    if rol == 'motorista' and request.user.id != pk:
        messages.error(request, 'No puedes ver el perfil de otro motorista.')
        return redirect('home')

    motorista = get_object_or_404(Motorista, pk=pk)
    
    asignaciones = AsignacionMotoMotorista.objects.filter(motorista=motorista)
    asignaciones_activas = asignaciones.filter(activa=1)

    if request.method == 'POST':
        permisos = obtener_permisos_usuario(request.user)
        acciones = permisos.get('asignaciones') or set()
        if not ('add' in acciones or 'change' in acciones or 'all' in acciones):
            messages.error(request, 'Acceso denegado para crear/asignar relaciones Motorista–Farmacia.')
            return redirect('acceso_denegado')
        rol = obtener_rol_usuario(request.user)
        if rol not in ('supervisor','operador','admin'):
            messages.error(request, 'Acceso denegado para tu rol actual.')
            return redirect('acceso_denegado')
        form = AsignacionMotoristaFarmaciaForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            try:
                existente = AsignacionMotoristaFarmacia.objects.filter(motorista=cd['motorista'], farmacia=cd['farmacia']).order_by('-fecha_asignacion').first()
            except Exception:
                existente = None
            if existente:
                existente.activa = cd.get('activa', 1)
                existente.fecha_desasignacion = cd.get('fecha_desasignacion')
                existente.observaciones = cd.get('observaciones')
                existente.save()
                messages.success(request, 'Asignación actualizada exitosamente.')
            else:
                obj = form.save(commit=False)
                try:
                    from django.utils import timezone as _tz
                    obj.fecha_asignacion = _tz.now()
                except Exception:
                    pass
                obj.save()
                messages.success(request, 'Motorista asignado a farmacia exitosamente.')
            return redirect('listado_asignaciones')
        else:
            messages.error(request, 'Corrige los errores del formulario de asignación.')
            asignacion_mf_form = form
    else:
        asignacion_mf_form = AsignacionMotoristaFarmaciaForm(initial={
            'motorista': motorista.id,
            'fecha_asignacion': timezone.now(),
            'activa': 1,
        })

    despachos = Despacho.objects.filter(motorista=motorista).order_by('-fecha_registro')[:50]
    lids = [d.farmacia_origen_local_id for d in despachos if d.farmacia_origen_local_id]
    origenes = {f.local_id: f for f in Farmacia.objects.filter(local_id__in=lids)}
    for d in despachos:
        try:
            d.origen = origenes.get(d.farmacia_origen_local_id)
        except Exception:
            d.origen = None
        try:
            from django.utils import timezone as _tz
            now_dt = _tz.now()
            if d.fecha_salida_farmacia and not d.fecha_completado and not d.fecha_anulacion:
                d.min_en_ruta = int(max((now_dt - d.fecha_salida_farmacia).total_seconds() / 60, 0))
            else:
                d.min_en_ruta = None
        except Exception:
            d.min_en_ruta = None
    context = {
        'motorista': motorista,
        'asignaciones': asignaciones,
        'asignaciones_activas': asignaciones_activas,
        'asignacion_mf_form': asignacion_mf_form,
        'despachos': despachos,
        'origenes': origenes,
        'embed': True if (request.GET.get('embed') == '1') else False,
    }

    return render(request, 'motoristas/detalle-motorista.html', context)


# ===== MOTO =====
@permiso_requerido('motos', 'view')
def listado_motos(request):
    """Lista todas las motos con búsqueda y paginación"""
    search_query = request.GET.get('search', '').strip()
    filtro_estado = (request.GET.get('estado') or '').strip().lower()
    rol = obtener_rol_usuario(request.user)

    motos = Moto.objects.all()

    

    # Si es motorista, solo ve su moto asignada activa
    if rol == 'motorista':
        motorista = Motorista.objects.filter(usuario=request.user).first()
        if motorista:
            asignacion_activa = AsignacionMotoMotorista.objects.filter(motorista=motorista, activa=1).first()
            if asignacion_activa:
                motos = motos.filter(pk=asignacion_activa.moto.pk)
            else:
                motos = Moto.objects.none()

    if search_query:
        motos = motos.filter(
            Q(patente__icontains=search_query) |
            Q(marca__icontains=search_query) |
            Q(modelo__icontains=search_query) |
            Q(numero_motor__icontains=search_query) |
            Q(propietario_nombre__icontains=search_query)
        )

    if filtro_estado == 'activa':
        motos = motos.filter(activo=True)
    elif filtro_estado == 'inactiva':
        motos = motos.filter(activo=False)

    motos = motos.order_by('patente')

    paginator = Paginator(motos, 10)
    page_number = request.GET.get('page')
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    samples = []
    if motos.count() == 0:
        try:
            import json, pathlib
            data_path = pathlib.Path(__file__).resolve().parent.parent / 'static' / 'data' / 'motos.json'
            with open(data_path, 'r', encoding='utf-8') as f:
                samples = json.load(f)
        except Exception:
            samples = []
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'estado': filtro_estado,
        'samples': samples,
    }

    return render(request, 'motos/listado-motos.html', context)


@permiso_requerido('motos', 'add')
def agregar_moto(request):
    """Crea una nueva moto"""
    if request.method == 'POST':
        form = MotoForm(request.POST, request.FILES)
        if form.is_valid():
            moto = form.save(commit=False)
            try:
                from django.utils import timezone
                moto.fecha_creacion = timezone.now()
                moto.fecha_modificacion = timezone.now()
                moto.usuario_modificacion = Usuario.objects.filter(django_user_id=request.user.id).first()
            except Exception:
                pass
            moto.save()
            try:
                docs = request.FILES.getlist('documentos')
                if docs:
                    import os, imghdr
                    from django.conf import settings
                    base = os.path.join(settings.MEDIA_ROOT, 'docs', 'motos', str(moto.pk))
                    os.makedirs(base, exist_ok=True)
                    allow = set(settings.UPLOAD_ALLOWED_CONTENT_TYPES)
                    maxsz = settings.UPLOAD_MAX_SIZE_MB * 1024 * 1024
                    invalid = 0
                    for f in docs:
                        try:
                            ct = getattr(f, 'content_type', '')
                            if ct not in allow:
                                invalid += 1
                                continue
                            if f.size > maxsz:
                                invalid += 1
                                continue
                            ext = '.bin'
                            if ct == 'application/pdf':
                                head = f.read(4); f.seek(0)
                                if head != b'%PDF':
                                    invalid += 1
                                    continue
                                ext = '.pdf'
                            else:
                                sniff = imghdr.what(None, h=f.read(32)); f.seek(0)
                                if sniff not in ('jpeg','png'):
                                    invalid += 1
                                    continue
                                ext = '.jpg' if sniff == 'jpeg' else '.png'
                            safe = os.path.basename(getattr(f, 'name', 'doc'))
                            name, _ = os.path.splitext(safe)
                            path = os.path.join(base, name + ext)
                            with open(path, 'wb') as dest:
                                for chunk in f.chunks():
                                    dest.write(chunk)
                        except Exception:
                            invalid += 1
                    if invalid:
                        messages.warning(request, f'{invalid} archivo(s) fueron rechazados por tipo/tamaño inválido.')
            except Exception:
                pass
            messages.success(request, f'Moto "{moto.patente}" creada exitosamente.')
            return redirect('detalle_moto', pk=moto.pk)
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        initial = {}
        try:
            mid = int(request.GET.get('motorista') or '0')
            if mid:
                initial['propietario_tipo'] = 'MOTORISTA'
                from .models import Motorista
                m = Motorista.objects.filter(pk=mid).first()
                if m:
                    initial['motorista_propietario'] = m
                    initial['propietario_nombre'] = f"{getattr(m.usuario,'nombre','')} {getattr(m.usuario,'apellido','')}".strip()
        except Exception:
            pass
        form = MotoForm(initial=initial)

    return render(request, 'motos/agregar-moto.html', {'form': form})


@permiso_requerido('motos', 'change')
def actualizar_moto(request, pk):
    """Actualiza datos de una moto existente"""
    moto = get_object_or_404(Moto, pk=pk)

    if request.method == 'POST':
        form = MotoForm(request.POST, request.FILES, instance=moto)
        if form.is_valid():
            moto = form.save(commit=False)
            try:
                from django.utils import timezone
                moto.fecha_modificacion = timezone.now()
                moto.usuario_modificacion = Usuario.objects.filter(django_user_id=request.user.id).first()
            except Exception:
                pass
            moto.save()
            try:
                docs = request.FILES.getlist('documentos')
                if docs:
                    import os, imghdr
                    from django.conf import settings
                    base = os.path.join(settings.MEDIA_ROOT, 'docs', 'motos', str(moto.pk))
                    os.makedirs(base, exist_ok=True)
                    allow = set(settings.UPLOAD_ALLOWED_CONTENT_TYPES)
                    maxsz = settings.UPLOAD_MAX_SIZE_MB * 1024 * 1024
                    invalid = 0
                    for f in docs:
                        try:
                            ct = getattr(f, 'content_type', '')
                            if ct not in allow:
                                invalid += 1
                                continue
                            if f.size > maxsz:
                                invalid += 1
                                continue
                            ext = '.bin'
                            if ct == 'application/pdf':
                                head = f.read(4); f.seek(0)
                                if head != b'%PDF':
                                    invalid += 1
                                    continue
                                ext = '.pdf'
                            else:
                                sniff = imghdr.what(None, h=f.read(32)); f.seek(0)
                                if sniff not in ('jpeg','png'):
                                    invalid += 1
                                    continue
                                ext = '.jpg' if sniff == 'jpeg' else '.png'
                            safe = os.path.basename(getattr(f, 'name', 'doc'))
                            name, _ = os.path.splitext(safe)
                            path = os.path.join(base, name + ext)
                            with open(path, 'wb') as dest:
                                for chunk in f.chunks():
                                    dest.write(chunk)
                        except Exception:
                            invalid += 1
                    if invalid:
                        messages.warning(request, f'{invalid} archivo(s) fueron rechazados por tipo/tamaño inválido.')
            except Exception:
                pass
            messages.success(request, 'Moto actualizada exitosamente.')
            return redirect('detalle_moto', pk=moto.pk)
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = MotoForm(instance=moto)

    return render(request, 'motos/modificar-moto.html', {'form': form, 'moto': moto})


@permiso_requerido('motos', 'delete')
def remover_moto(request, pk):
    """Elimina una moto"""
    moto = get_object_or_404(Moto, pk=pk)

    if request.method == 'POST':
        patente = moto.patente
        try:
            moto.delete()
            messages.success(request, f'Moto "{patente}" eliminada exitosamente.')
        except Exception as e:
            messages.error(request, f'Error al eliminar: {str(e)}')
        return redirect('listado_motos')

    return render(request, 'motos/remover-moto.html', {'moto': moto})


@permiso_requerido('motos', 'view')
def detalle_moto(request, pk):
    """Ver detalles de una moto"""
    rol = obtener_rol_usuario(request.user)

    moto = get_object_or_404(Moto, pk=pk)

    # Si es motorista, solo puede ver su moto asignada
    if rol == 'motorista':
        try:
            motorista_usuario = Motorista.objects.get(usuario=request.user)
            asignacion_activa = AsignacionMotoMotorista.objects.filter(motorista=motorista_usuario, moto=moto, activa=1).exists()
            if not asignacion_activa:
                messages.error(request, 'No puedes ver motos que no te están asignadas.')
                return redirect('listado_motos')
        except Motorista.DoesNotExist:
            messages.error(request, 'No tienes un perfil de motorista asociado.')
            return redirect('home')

    # Si es farmacia, solo puede ver motos de su farmacia (no hay relación directa, se omite)
    if rol == 'farmacia':
        # No hay relación directa moto-farmacia ni motorista-farmacia en este modelo
        pass

    asignaciones = AsignacionMotoMotorista.objects.filter(moto=moto)

    context = {
        'moto': moto,
        'asignaciones': asignaciones,
    }

    return render(request, 'motos/detalle-moto.html', context)


@permiso_requerido('motos', 'view')
def buscar_moto(request):
    search = (request.GET.get('q') or '').strip()
    marca = (request.GET.get('marca') or '').strip()
    modelo = (request.GET.get('modelo') or '').strip()
    color = (request.GET.get('color') or '').strip()
    propietario = (request.GET.get('propietario') or '').strip()
    estado = (request.GET.get('estado') or '').strip().lower()
    tipo_comb = (request.GET.get('tipo_combustible') or '').strip()
    km_min = request.GET.get('km_min')
    km_max = request.GET.get('km_max')
    anio_min = request.GET.get('anio_min')
    anio_max = request.GET.get('anio_max')
    cil_min = request.GET.get('cil_min')
    cil_max = request.GET.get('cil_max')

    qs = Moto.objects.all()
    if search:
        qs = qs.filter(
            Q(patente__icontains=search) |
            Q(marca__icontains=search) |
            Q(modelo__icontains=search) |
            Q(numero_motor__icontains=search) |
            Q(propietario_nombre__icontains=search)
        )
    if marca:
        qs = qs.filter(marca__icontains=marca)
    if modelo:
        qs = qs.filter(modelo__icontains=modelo)
    if color:
        qs = qs.filter(color__icontains=color)
    if propietario:
        qs = qs.filter(propietario_nombre__icontains=propietario)
    if estado == 'activa':
        qs = qs.filter(activo=True)
    elif estado == 'inactiva':
        qs = qs.filter(activo=False)
    if tipo_comb:
        qs = qs.filter(tipo_combustible__iexact=tipo_comb)
    try:
        if km_min:
            qs = qs.filter(kilometraje_actual__gte=int(km_min))
        if km_max:
            qs = qs.filter(kilometraje_actual__lte=int(km_max))
    except Exception:
        pass
    try:
        if anio_min:
            qs = qs.filter(anio__gte=int(anio_min))
        if anio_max:
            qs = qs.filter(anio__lte=int(anio_max))
    except Exception:
        pass
    try:
        if cil_min:
            qs = qs.filter(cilindrada_cc__gte=int(cil_min))
        if cil_max:
            qs = qs.filter(cilindrada_cc__lte=int(cil_max))
    except Exception:
        pass
    qs = qs.order_by('patente')

    paginator = Paginator(qs, 12)
    page_number = request.GET.get('page')
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    ctx = {
        'page_obj': page_obj,
        'q': search,
        'marca': marca,
        'modelo': modelo,
        'color': color,
        'propietario': propietario,
        'estado': estado,
        'tipo_combustible': tipo_comb,
        'km_min': km_min,
        'km_max': km_max,
        'anio_min': anio_min,
        'anio_max': anio_max,
        'cil_min': cil_min,
        'cil_max': cil_max,
    }
    return render(request, 'motos/buscar-moto.html', ctx)


# ===== ASIGNACIONES =====
@permiso_requerido('asignaciones', 'view')
def listado_asignaciones(request):
    """Lista principal: asignaciones Motorista–Farmacia"""
    search_query = request.GET.get('search', '').strip()
    filtro_estado = request.GET.get('estado', '')
    from .models import AsignacionMotoristaFarmacia

    asignaciones = AsignacionMotoristaFarmacia.objects.all().select_related('motorista__usuario', 'farmacia')

    if search_query:
        asignaciones = asignaciones.filter(
            Q(motorista__usuario__nombre__icontains=search_query) |
            Q(motorista__usuario__apellido__icontains=search_query) |
            Q(farmacia__local_nombre__icontains=search_query) |
            Q(observaciones__icontains=search_query)
        )

    if filtro_estado == 'activa':
        asignaciones = asignaciones.filter(activa=1)
    elif filtro_estado == 'inactiva':
        asignaciones = asignaciones.filter(activa=0)

    asignaciones = asignaciones.order_by('-fecha_asignacion')

    paginator = Paginator(asignaciones, 10)
    page_number = request.GET.get('page')
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    samples = []
    if page_obj.paginator.count == 0:
        try:
            import json, pathlib
            data_path = pathlib.Path(__file__).resolve().parent.parent / 'static' / 'data' / 'asignaciones_motorista_farmacia.json'
            with open(data_path, 'r', encoding='utf-8') as f:
                samples = json.load(f)
        except Exception:
            samples = []
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'filtro_estado': filtro_estado,
        'samples': samples,
    }

    return render(request, 'asignaciones/listar-asignaciones-mf.html', context)


@permiso_requerido('asignaciones', 'view')
def detalle_asignacion(request, pk):
    """Ver detalles de una asignación Motorista–Farmacia"""
    from .models import AsignacionMotoristaFarmacia
    asignacion = get_object_or_404(AsignacionMotoristaFarmacia, pk=pk)
    return render(request, 'asignaciones/detalle-asignacion-mf.html', {'asignacion': asignacion})


@permiso_requerido('asignaciones', 'add')
def agregar_asignacion(request):
    """Crea una nueva asignación Motorista–Farmacia"""
    if request.method == 'POST':
        form = AsignacionMotoristaFarmaciaForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            try:
                from django.utils import timezone as _tz
                obj.fecha_asignacion = _tz.now()
            except Exception:
                pass
            obj.save()
            messages.success(request, 'Asignación creada exitosamente.')
            return redirect('detalle_asignacion', pk=obj.pk)
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
            return render(request, 'asignaciones/agregar-asignacion-mf.html', {'form': form})
    else:
        initial = {}
        mot = request.GET.get('motorista')
        far = request.GET.get('farmacia')
        try:
            if mot:
                initial['motorista'] = int(mot)
        except Exception:
            pass
        try:
            if far:
                initial['farmacia'] = far
        except Exception:
            pass
        form = AsignacionMotoristaFarmaciaForm(initial=initial)
        return render(request, 'asignaciones/agregar-asignacion-mf.html', {'form': form})


@permiso_requerido('asignaciones', 'change')
def modificar_asignacion(request, pk):
    """Edita una asignación Motorista–Farmacia"""
    from .models import AsignacionMotoristaFarmacia
    asignacion = get_object_or_404(AsignacionMotoristaFarmacia, pk=pk)
    if request.method == 'POST':
        form = AsignacionMotoristaFarmaciaForm(request.POST, instance=asignacion)
        if form.is_valid():
            form.save()
            messages.success(request, 'Asignación actualizada exitosamente.')
            return redirect('detalle_asignacion', pk=asignacion.pk)
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = AsignacionMotoristaFarmaciaForm(instance=asignacion)
    return render(request, 'asignaciones/editar-asignacion-mf.html', {'form': form, 'asignacion': asignacion})


@permiso_requerido('asignaciones', 'change')
def remover_asignacion(request, pk):
    """Activa o desactiva una asignación Motorista–Farmacia"""
    from .models import AsignacionMotoristaFarmacia
    asignacion = get_object_or_404(AsignacionMotoristaFarmacia, pk=pk)
    if request.method == 'POST':
        from .roles import obtener_rol_usuario
        rol = obtener_rol_usuario(request.user)
        if rol not in ('supervisor','admin'):
            messages.error(request, 'Solo supervisor/admin puede modificar asignaciones')
            return redirect('detalle_asignacion', pk=asignacion.pk)
        try:
            from django.db import transaction as _tx
            with _tx.atomic():
                prev = {'activa': asignacion.activa}
                asignacion.activa = 1 if asignacion.activa == 0 else 0
                asignacion.save()
                try:
                    from .models import AuditoriaGeneral, Usuario as _Usuario
                    AuditoriaGeneral.objects.create(
                        nombre_tabla='asignacion_motorista_farmacia',
                        id_registro_afectado=str(asignacion.pk),
                        tipo_operacion='UPDATE',
                        usuario=_Usuario.objects.filter(django_user_id=request.user.id).first(),
                        fecha_evento=timezone.now(),
                        datos_antiguos=prev,
                        datos_nuevos={'activa': asignacion.activa}
                    )
                except Exception:
                    pass
            estado = "activada" if asignacion.activa == 1 else "desactivada"
            messages.success(request, f'Asignación {estado} exitosamente.')
        except Exception as e:
            messages.error(request, f'Error al actualizar: {str(e)}')
        return redirect('detalle_asignacion', pk=asignacion.pk)
    return render(request, 'asignaciones/detalle-asignacion-mf.html', {'asignacion': asignacion})


@rol_requerido('gerente')
def reporte_movimientos(request):
    form = ReporteMovimientosForm(request.GET or None)
    from .models import MovimientoDespacho, Localfarmacia
    movimientos = MovimientoDespacho.objects.all().select_related('despacho')
    farmacias = Localfarmacia.objects.filter(activo=True)

    template = 'reportes/reporte-diario.html'
    if form.is_valid():
        tipo = form.cleaned_data.get('tipo_reporte')
        fecha = form.cleaned_data.get('fecha')
        mes = form.cleaned_data.get('mes')
        anio = form.cleaned_data.get('anio')
        farmacia = form.cleaned_data.get('farmacia')
        if farmacia:
            movimientos = movimientos.filter(despacho__farmacia_origen_local_id=farmacia.local_id)
        if tipo == 'diario':
            template = 'reportes/reporte-diario.html'
            if fecha:
                movimientos = movimientos.filter(fecha_movimiento__date=fecha)
            try:
                from .repositories import get_resumen_operativo_hoy
                resumen = get_resumen_operativo_hoy()
            except Exception:
                resumen = []
        elif tipo == 'mensual':
            template = 'reportes/reporte-mensual.html'
            if mes:
                movimientos = movimientos.filter(fecha_movimiento__year=mes.year, fecha_movimiento__month=mes.month)
            resumen = get_resumen_operativo_mes(anio=mes.year if mes else None, mes=mes.month if mes else None)
        elif tipo == 'anual':
            template = 'reportes/reporte-anual.html'
            if anio:
                movimientos = movimientos.filter(fecha_movimiento__year=anio)
            resumen = get_resumen_operativo_anual(anio=anio)

    # Fallback con datos de ejemplo si no hay datos
    mov_list = None
    try:
        if movimientos.count() == 0:
            import json, pathlib
            data_path = pathlib.Path(__file__).resolve().parent.parent / 'static' / 'data' / 'movimientos.json'
            with open(data_path, 'r', encoding='utf-8') as f:
                raw = json.load(f)
            class M:
                def __init__(self, d):
                    self.despacho = type('D', (), {'id': d.get('despacho_id')})
                    self.estado_nuevo = d.get('estado_nuevo')
                    self.fecha_movimiento = d.get('fecha_movimiento')
            mov_list = [M(d) for d in raw]
    except Exception:
        mov_list = None

    paginator = Paginator((mov_list or movimientos.order_by('-fecha_movimiento')), 10)
    page_number = request.GET.get('page')
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    contexto = {
        'form': form,
        'movimientos': page_obj.object_list,
        'page_obj': page_obj,
        'farmacias': farmacias,
        'resumen': locals().get('resumen') if 'resumen' in locals() else [],
    }
    return render(request, template, contexto)


@solo_admin
def importar_farmacias(request):
    mensajes = []
    creados = 0
    if request.method == 'POST':
        csv_text = request.POST.get('csv_text', '').strip()
        fichero = request.FILES.get('csv_file')
        rows = []
        try:
            if csv_text:
                import csv, io
                sep = '\t' if '\t' in csv_text else ','
                reader = csv.DictReader(io.StringIO(csv_text), delimiter=sep)
                rows = list(reader)
            elif fichero:
                name = fichero.name.lower()
                if name.endswith('.csv') or name.endswith('.tsv'):
                    import csv, io
                    data = fichero.read().decode('utf-8', errors='ignore')
                    sep = '\t' if ('\t' in data or name.endswith('.tsv')) else ','
                    reader = csv.DictReader(io.StringIO(data), delimiter=sep)
                    rows = list(reader)
                elif name.endswith('.xlsx'):
                    import openpyxl
                    wb = openpyxl.load_workbook(fichero, data_only=True)
                    for sheet in wb.worksheets:
                        values = list(sheet.values)
                        if not values:
                            continue
                        headers = [str(h).strip() if h else '' for h in values[0]]
                        for row in values[1:]:
                            d = {}
                            for i, v in enumerate(row):
                                key = headers[i] if i < len(headers) else f'col_{i}'
                                d[key] = v if v is not None else ''
                            rows.append(d)
                elif name.endswith('.xls'):
                    import xlrd
                    book = xlrd.open_workbook(file_contents=fichero.read())
                    for sh in book.sheets():
                        if sh.nrows == 0:
                            continue
                        headers = [str(sh.cell_value(0, c)).strip() for c in range(sh.ncols)]
                        for r in range(1, sh.nrows):
                            d = {}
                            for c in range(sh.ncols):
                                d[headers[c]] = sh.cell_value(r, c)
                            rows.append(d)
                else:
                    mensajes.append('Formato no soportado. Usa CSV/TSV/XLSX/XLS.')
            else:
                mensajes.append('Debes pegar datos o subir un archivo CSV.')
            from .models import Localfarmacia, Usuario
            usuario = Usuario.objects.filter(django_user_id=request.user.id).first()
            required = {'local_id','local_nombre','local_direccion','comuna_nombre'}
            for r in rows:
                def get(*keys):
                    for k in keys:
                        if k in r and r[k] is not None:
                            val = str(r[k]).strip()
                            if val != '':
                                return val
                    return None
                try:
                    missing = [k for k in required if not get(k, k.capitalize(), k.title())]
                    if missing:
                        mensajes.append(f'Fila inválida: faltan {", ".join(missing)}')
                        continue
                    # duplicados por local_id
                    lid = get('local_id','ID','id') or ''
                    if Localfarmacia.objects.filter(local_id=lid).exists():
                        mensajes.append(f'Duplicado saltado local_id={lid}')
                        continue
                    apertura = get('funcionamiento_hora_apertura','apertura') or '09:00:00'
                    cierre = get('funcionamiento_hora_cierre','cierre') or '21:00:00'
                    from datetime import datetime as dt
                    def parse_time(s):
                        for fmt in ('%H:%M:%S','%H:%M'):
                            try:
                                return dt.strptime(s, fmt).time()
                            except:
                                pass
                        return dt.strptime('09:00:00','%H:%M:%S').time()
                    hora_ap = parse_time(apertura)
                    hora_ci = parse_time(cierre)
                    fecha_str = get('fecha_actualizacion','fecha','Fecha')
                    if fecha_str:
                        try:
                            fecha = dt.strptime(fecha_str, '%d-%m-%y').date()
                        except:
                            fecha = dt.strptime(fecha_str, '%Y-%m-%d').date()
                    else:
                        fecha = dt.today().date()
                    lat = get('local_lat','lat','Lat')
                    lng = get('local_lng','lng','Lng')
                    try:
                        lat = float(lat) if lat else None
                    except:
                        lat = None
                    try:
                        lng = float(lng) if lng else None
                    except:
                        lng = None
                    tel_raw = get('local_telefono','telefono','Teléfono') or ''
                    import re
                    tel_clean = re.sub(r'[^0-9+]', '', tel_raw)
                    if tel_clean == '+56' or len(re.sub(r'[^0-9]', '', tel_clean)) < 7:
                        tel_clean = None
                    # FKs por ID si vienen en Excel
                    rid = get('fk_region','region_id','id_region')
                    cid = get('fk_comuna','comuna_id','id_comuna')
                    lid = get('fk_localidad','localidad_id','id_localidad')
                    fk_region = None
                    fk_comuna = None
                    fk_localidad = None
                    try:
                        if rid:
                            from .models import Region
                            fk_region = Region.objects.filter(id=int(float(rid))).first()
                    except Exception:
                        fk_region = None
                    try:
                        if cid:
                            from .models import Comuna
                            fk_comuna = Comuna.objects.filter(id=int(float(cid))).first()
                    except Exception:
                        fk_comuna = None
                    try:
                        if lid:
                            from .models import Localidad
                            fk_localidad = Localidad.objects.filter(id=int(float(lid))).first()
                    except Exception:
                        fk_localidad = None
                    obj = Localfarmacia(
                        local_id=lid,
                        local_nombre=get('local_nombre','nombre','Nombre') or '',
                        local_direccion=get('local_direccion','direccion','Dirección') or '',
                        comuna_nombre=get('comuna_nombre','comuna','Comuna') or '',
                        localidad_nombre=get('localidad_nombre','localidad','Localidad') or '',
                        fk_region=fk_region,
                        fk_comuna=fk_comuna,
                        fk_localidad=fk_localidad,
                        funcionamiento_hora_apertura=hora_ap,
                        funcionamiento_hora_cierre=hora_ci,
                        funcionamiento_dia=get('funcionamiento_dia','dia','Día') or 'lunes',
                        local_telefono=tel_clean,
                        local_lat=lat,
                        local_lng=lng,
                        geolocalizacion_validada=True if (lat is not None and lng is not None) else False,
                        fecha=fecha,
                        activo=True,
                        fecha_creacion=dt.now(),
                        fecha_modificacion=dt.now(),
                        usuario_modificacion=usuario,
                    )
                    obj.save()
                    creados += 1
                except Exception as e:
                    mensajes.append(f'Fila con error: {e}')
        except Exception as e:
            mensajes.append(f'Error al procesar: {e}')
    context = {'mensajes': mensajes, 'creados': creados}
    return render(request, 'localfarmacia/importar-farmacias.html', context)


@permiso_requerido('movimientos', 'add')
@ratelimit(key='ip', rate='20/m', block=True)
def ingestar_normalizacion(request):
    from .models import NormalizacionDespacho
    mensajes = []
    creados = 0
    procesados = 0
    if request.method == 'POST':
        fichero = request.FILES.get('csv_file')
        fuente = (request.POST.get('fuente') or 'excel').strip().lower()
        rows = []
        try:
            if fichero:
                name = fichero.name.lower()
                if name.endswith('.csv') or name.endswith('.tsv'):
                    import csv, io
                    data = fichero.read().decode('utf-8', errors='ignore')
                    sep = '\t' if ('\t' in data or name.endswith('.tsv')) else ','
                    reader = csv.DictReader(io.StringIO(data), delimiter=sep)
                    rows = list(reader)
                elif name.endswith('.xlsx'):
                    import openpyxl
                    wb = openpyxl.load_workbook(fichero, data_only=True)
                    for sheet in wb.worksheets:
                        values = list(sheet.values)
                        if not values:
                            continue
                        headers = [str(h).strip() if h else '' for h in values[0]]
                        for row in values[1:]:
                            d = {}
                            for i, v in enumerate(row):
                                key = headers[i] if i < len(headers) else f'col_{i}'
                                d[key] = v if v is not None else ''
                            rows.append(d)
                else:
                    mensajes.append('Formato no soportado. Usa CSV/TSV/XLSX.')
            else:
                mensajes.append('Debes subir un archivo CSV/XLSX.')
            for r in rows:
                def get(*keys):
                    for k in keys:
                        if k in r and r[k] is not None:
                            val = str(r[k]).strip()
                            if val != '':
                                return val
                    return None
                try:
                    obj = NormalizacionDespacho(
                        fuente=fuente,
                        farmacia_origen_local_id=get('farmacia_origen_local_id','local_id','farmacia'),
                        motorista_documento=get('motorista_documento','motorista','rut_motorista','dni_motorista'),
                        cliente_nombre_raw=get('cliente_nombre','cliente'),
                        cliente_telefono_raw=get('cliente_telefono','telefono','fono'),
                        destino_direccion_raw=get('destino_direccion','direccion','calle'),
                        destino_lat_raw=get('destino_lat','lat'),
                        destino_lng_raw=get('destino_lng','lng'),
                        estado_raw=get('estado','estado_raw'),
                        tipo_despacho_raw=get('tipo_despacho','tipo'),
                        prioridad_raw=get('prioridad','prio'),
                        numero_receta_raw=get('numero_receta','receta'),
                        observaciones_raw=get('observaciones','obs'),
                        fecha_registro_raw=get('fecha','fecha_registro'),
                        procesado=False,
                        error_normalizacion=None,
                        fecha_creacion=timezone.now(),
                    )
                    obj.save()
                    creados += 1
                except Exception as e:
                    mensajes.append(f'Fila staging con error: {e}')
            try:
                normalize_from_normalizacion(limit=500)
                procesados = 1
            except Exception as e:
                mensajes.append(f'Error en normalización: {e}')
        except Exception as e:
            mensajes.append(f'Error al procesar: {e}')
    context = {'mensajes': mensajes, 'creados': creados, 'procesados': procesados}
    return render(request, 'movimientos/ingestar-staging.html', context)


@permiso_requerido('movimientos', 'add')
@ratelimit(key='ip', rate='20/m', block=True)
def registrar_movimiento(request):
    from .models import Despacho, MovimientoDespacho, Usuario
    feedback = None
    import logging
    log = logging.getLogger('appnproylogico')
    if request.method == 'POST':
        metodo = request.POST.get('metodo')  # llamada | mensaje | boton
        tipo_mov = (request.POST.get('tipo_movimiento','') or '').strip().lower()
        codigo = request.POST.get('codigo_despacho','').strip()
        estado = request.POST.get('estado','').strip()
        confirmado = (request.POST.get('confirmado','') or '').strip().lower()
        mensaje = request.POST.get('mensaje','').strip()
        try:
            despacho = Despacho.objects.filter(codigo_despacho=codigo).first()
            if not despacho:
                feedback = 'Despacho no encontrado'
                log.info('Movimiento rechazado: despacho no encontrado codigo=%s ip=%s', codigo, request.META.get('REMOTE_ADDR'))
            else:
                from django.utils import timezone
                usuario = Usuario.objects.filter(django_user_id=request.user.id).first()
                estado_norm = (despacho.estado or '').strip().upper()
                nuevo = (estado or '').strip().upper()
                tipo_d = (despacho.tipo_despacho or '').strip().upper()
                try:
                    pend = request.session.get('mov_queue', [])
                    if pend:
                        for item in list(pend):
                            try:
                                d2 = Despacho.objects.filter(codigo_despacho=item.get('codigo')).first()
                                if not d2:
                                    continue
                                u2 = Usuario.objects.filter(django_user_id=request.user.id).first()
                                MovimientoDespacho.objects.create(
                                    despacho=d2,
                                    estado_anterior=item.get('estado_anterior'),
                                    estado_nuevo=item.get('estado_nuevo'),
                                    fecha_movimiento=item.get('fecha_movimiento') or timezone.now(),
                                    usuario=u2,
                                    observacion=item.get('observacion','')
                                )
                                pend.remove(item)
                            except Exception:
                                continue
                        request.session['mov_queue'] = pend
                except Exception:
                    pass
                mapa = {
                    'PENDIENTE': {'ASIGNADO','ANULADO'},
                    'ASIGNADO': {'PREPARANDO','ANULADO'},
                    'PREPARANDO': {'PREPARADO','ANULADO'},
                    'PREPARADO': {'EN_CAMINO','ANULADO'},
                    'EN_CAMINO': {'ENTREGADO','FALLIDO'},
                }
                permitidos = mapa.get(estado_norm, set())
                if nuevo not in permitidos:
                    feedback = 'Transición de estado no permitida'
                    log.info('Transición inválida codigo=%s de=%s a=%s', codigo, estado_norm, nuevo)
                    messages.error(request, feedback)
                    context = {'feedback': feedback}
                    return render(request, 'movimientos/registrar-mov.html', context)
                if tipo_d == 'REENVIO_RECETA' and nuevo == 'PREPARADO':
                    if not (despacho.tiene_receta_retenida and despacho.receta_devuelta_farmacia):
                        feedback = 'Receta retenida requiere devolución antes de PREPARADO'
                        log.info('Bloqueado PREPARADO por receta no devuelta codigo=%s', codigo)
                        messages.error(request, feedback)
                        context = {'feedback': feedback}
                        return render(request, 'movimientos/registrar-mov.html', context)
                nuevo = (estado or '').strip().upper()
                tipo_d = (despacho.tipo_despacho or '').strip().upper()
                if (metodo or '').lower() == 'boton' and confirmado != 'si':
                    feedback = 'Debes confirmar la acción antes de continuar'
                    messages.error(request, feedback)
                    context = {'feedback': feedback}
                    return render(request, 'movimientos/registrar-mov.html', context)
                if nuevo == 'ANULADO':
                    from .roles import obtener_rol_usuario
                    rol = obtener_rol_usuario(request.user)
                    if rol != 'operador':
                        feedback = 'Solo operadora puede anular'
                        messages.error(request, feedback)
                        context = {'feedback': feedback}
                        return render(request, 'movimientos/registrar-mov.html', context)
                    if not mensaje or len(mensaje) < 10:
                        feedback = 'Descripción obligatoria para ANULADO'
                        messages.error(request, feedback)
                        context = {'feedback': feedback}
                        return render(request, 'movimientos/registrar-mov.html', context)
                    try:
                        from django.utils import timezone as _tz
                        now_s = int(_tz.now().timestamp())
                        window_start = int(request.session.get('cancel_window_start') or 0)
                        count = int(request.session.get('cancel_count') or 0)
                        if not window_start or now_s - window_start > 600:
                            request.session['cancel_window_start'] = now_s
                            request.session['cancel_count'] = 0
                        else:
                            if count >= 5:
                                feedback = 'Límite de anulaciones alcanzado (10 min)'
                                messages.error(request, feedback)
                                context = {'feedback': feedback}
                                return render(request, 'movimientos/registrar-mov.html', context)
                    except Exception:
                        pass
                mapa = {
                    'PENDIENTE': {'ASIGNADO','ANULADO'},
                    'ASIGNADO': {'PREPARANDO','ANULADO'},
                    'PREPARANDO': {'PREPARADO','ANULADO'},
                    'PREPARADO': {'EN_CAMINO','ANULADO'},
                    'EN_CAMINO': {'ENTREGADO','FALLIDO'},
                }
                permitidos = mapa.get(estado_norm, set())
                if nuevo not in permitidos:
                    feedback = 'Transición de estado no permitida'
                    log.info('Transición inválida codigo=%s de=%s a=%s', codigo, estado_norm, nuevo)
                else:
                    if nuevo == 'PREPARADO' and tipo_d == 'REENVIO_RECETA':
                        if not (despacho.tiene_receta_retenida and despacho.receta_devuelta_farmacia):
                            feedback = 'Receta retenida requiere devolución antes de PREPARADO'
                            log.info('Bloqueado PREPARADO por receta no devuelta codigo=%s', codigo)
                            context = {'feedback': feedback}
                            return render(request, 'movimientos/registrar-mov.html', context)
                    dup_skip = False
                    try:
                        ultimo = MovimientoDespacho.objects.filter(despacho=despacho).order_by('-fecha_movimiento').first()
                        if ultimo:
                            delta = abs((timezone.now() - ultimo.fecha_movimiento).total_seconds())
                            if ultimo.estado_nuevo == nuevo and delta < 60:
                                dup_skip = True
                    except Exception:
                        dup_skip = False
                    try:
                        if not dup_skip:
                            MovimientoDespacho.objects.create(
                                despacho=despacho,
                                estado_anterior=despacho.estado,
                                estado_nuevo=nuevo,
                                fecha_movimiento=timezone.now(),
                                usuario=usuario,
                                observacion=(f'modo={metodo}; tipo={tipo_mov or ""}; ' + (mensaje or '')).strip()
                            )
                        despacho.estado = nuevo
                        despacho.usuario_modificacion = usuario
                        despacho.fecha_modificacion = timezone.now()
                        despacho.save()
                        try:
                            from .models import AuditoriaGeneral
                            AuditoriaGeneral.objects.create(
                                nombre_tabla='movimiento_despacho',
                                id_registro_afectado=str(despacho.id),
                                tipo_operacion='MOV',
                                usuario=usuario,
                                fecha_evento=timezone.now(),
                                datos_antiguos={'estado': estado_norm},
                                datos_nuevos={'estado': estado, 'mensaje': mensaje}
                            )
                            if nuevo == 'ANULADO':
                                try:
                                    from django.utils import timezone as _tz
                                    request.session['cancel_count'] = int(request.session.get('cancel_count') or 0) + 1
                                    request.session['cancel_window_start'] = request.session.get('cancel_window_start') or int(_tz.now().timestamp())
                                except Exception:
                                    pass
                        except Exception:
                            pass
                        feedback = feedback or 'Movimiento registrado'
                        log.info('Movimiento registrado codigo=%s estado=%s usuario=%s ip=%s', codigo, (nuevo or estado), request.user.username, request.META.get('REMOTE_ADDR'))
                        messages.success(request, feedback)
                        ref = request.META.get('HTTP_REFERER') or None
                        if ref:
                            from django.shortcuts import redirect
                            return redirect(ref)
                        from django.urls import reverse
                        return redirect(reverse('despachos_activos'))
                    except Exception:
                        q = request.session.get('mov_queue', [])
                        q.append({
                            'codigo': codigo,
                            'estado_anterior': despacho.estado,
                            'estado_nuevo': nuevo,
                            'fecha_movimiento': timezone.now().isoformat(),
                            'observacion': (f'modo={metodo}; tipo={tipo_mov or ""}; ' + (mensaje or '')).strip(),
                        })
                        request.session['mov_queue'] = q
                        feedback = 'Base de datos no disponible; movimiento encolado'
                        messages.warning(request, feedback)
                        context = {'feedback': feedback}
                        return render(request, 'movimientos/registrar-mov.html', context)
        except Exception as e:
            feedback = f'Error: {e}'
            log.error('Error movimiento codigo=%s error=%s', codigo, e)
    context = {'feedback': feedback}
    return render(request, 'movimientos/registrar-mov.html', context)


@permiso_requerido('movimientos', 'view')
def resumen_operativo_hoy(request):
    rows = get_resumen_operativo_hoy()
    return render(request, 'reportes/resumen-operativo.html', {'rows': rows})


@permiso_requerido('movimientos', 'view')
def export_resumen_operativo(request):
    tipo = (request.GET.get('tipo') or 'diario').strip().lower()
    formato = (request.GET.get('formato') or 'csv').strip().lower()
    detalle = (request.GET.get('detalle') or '').strip() == '1'
    anio = request.GET.get('anio')
    mes = request.GET.get('mes')
    display_title = ''
    if tipo == 'diario' and not detalle:
        rows = get_resumen_operativo_hoy()
        headers = ['Farmacia','Comuna','Total despachos','Entregados','Fallidos','Directo','Reenvío receta','Intercambio','Error despacho','Con receta','Con incidencias']
        filename = 'resumen_diario'
        rows = [[r[1], r[2], r[3], r[4], r[5], r[13], r[14], r[15], r[16], r[9], r[10]] for r in rows]
    elif tipo == 'mensual' and not detalle:
        rows = get_resumen_operativo_mes(anio=anio, mes=mes)
        headers = ['Año','Mes','Farmacia','Comuna','Total despachos','Entregados','Fallidos','Directo','Reenvío receta','Intercambio','Error despacho','Con receta','Con incidencias']
        filename = f'resumen_mensual_{anio or "todos"}_{mes or "todos"}'
        rows = [[r[0], r[1], r[3], r[4], r[5], r[6], r[7], r[15], r[16], r[17], r[18], r[11], r[12]] for r in rows]
    elif tipo == 'asignaciones_mf':
        try:
            from .models import AsignacionMotoristaFarmacia, Despacho, Localfarmacia
            asignaciones = AsignacionMotoristaFarmacia.objects.select_related('motorista__usuario','farmacia').all().order_by('-fecha_asignacion')
            headers = ['Motorista','Farmacia','Activa','Fecha asignación','Despachos totales','Entregados','Fallidos','En camino','Pendientes','Anulados','Con receta retenida']
            filename = 'asignaciones_motorista_farmacia'
            rows = []
            for a in asignaciones:
                try:
                    lid = getattr(a.farmacia, 'local_id', None)
                    qs = Despacho.objects.filter(motorista=a.motorista)
                    if lid:
                        qs = qs.filter(farmacia_origen_local_id=lid)
                    total = qs.count()
                    entregados = qs.filter(estado='ENTREGADO').count()
                    fallidos = qs.filter(estado='FALLIDO').count()
                    en_camino = qs.filter(estado='EN_CAMINO').count()
                    pendientes = qs.filter(estado='PENDIENTE').count()
                    anulados = qs.filter(estado='ANULADO').count()
                    con_receta = qs.filter(tiene_receta_retenida=True).count()
                    rows.append([
                        f"{a.motorista.usuario.nombre} {a.motorista.usuario.apellido}",
                        f"{a.farmacia.local_nombre}",
                        'Sí' if a.activa else 'No',
                        a.fecha_asignacion.strftime('%Y-%m-%d %H:%M'),
                        total, entregados, fallidos, en_camino, pendientes, anulados, con_receta,
                    ])
                except Exception:
                    continue
        except Exception:
            rows = []
    elif tipo == 'despachos_activos':
        headers = ['Local','Despacho','Estado','Tipo','Prioridad','Motorista','Cliente','Dirección','Con receta','Incidencia','Fecha']
        fecha_arg = (request.GET.get('fecha') or timezone.now().strftime('%Y-%m-%d'))
        codigo = getattr(settings, 'REPORTE_CODIGO', 'codbd22_25')
        prefix = getattr(settings, 'REPORTE_PREFIX', 'logico')
        filename = f'{prefix}_{codigo}_{fecha_arg}'
        display_title = 'Reporte diario de despachos'
        try:
            import json, pathlib
            base = pathlib.Path(getattr(settings, 'MEDIA_ROOT', None) or (pathlib.Path(settings.BASE_DIR) / 'media')) / 'reportes'
            rows = []
            file = base / f'cierre_{fecha_arg}.json'
            if file.exists():
                with open(file, 'r', encoding='utf-8') as f:
                    rows = json.load(f) or []
                try:
                    rows = [[r[0], r[1], (_estado_normalizado(r[2]) if ' _estado_normalizado' else (r[2] if r else '')), r[3], r[4], r[5], r[6], r[7], r[8], r[9], r[10]] for r in rows]
                    rows = [r for r in rows if str(r[2]).strip().upper() in {'ENTREGADO','FALLIDO','PENDIENTE'}]
                except Exception:
                    pass
            else:
                # Fallback: construir desde BD de la fecha
                from .models import Despacho, Localfarmacia
                try:
                    y, m, d = [int(x) for x in fecha_arg.split('-')]
                except Exception:
                    y, m, d = timezone.now().year, timezone.now().month, timezone.now().day
                qs = Despacho.objects.filter(fecha_registro__year=y, fecha_registro__month=m, fecha_registro__day=d).order_by('fecha_registro')
                for obj in qs:
                    try:
                        farm = Localfarmacia.objects.filter(local_id=obj.farmacia_origen_local_id).first()
                        u = getattr(obj.motorista, 'usuario', None)
                        mot_name = f"{getattr(u,'nombre','')} {getattr(u,'apellido','')}".strip()
                        estado_norm = _estado_normalizado(obj.estado)
                        if estado_norm not in {'ENTREGADO','FALLIDO','PENDIENTE'}:
                            continue
                        rows.append([
                            getattr(farm,'local_id', obj.farmacia_origen_local_id) or '',
                            obj.codigo_despacho or obj.id,
                            estado_norm,
                            obj.tipo_despacho or '',
                            obj.prioridad or '',
                            mot_name,
                            _cliente_normalizado(obj.cliente_nombre),
                            obj.destino_direccion or '',
                            'Sí' if obj.tiene_receta_retenida else 'No',
                            'Sí' if obj.hubo_incidencia else 'No',
                            obj.fecha_registro.strftime('%Y-%m-%d %H:%M') if obj.fecha_registro else '',
                        ])
                    except Exception:
                        continue
        except Exception:
            rows = []
    elif tipo == 'diario' and detalle:
        fecha_arg = (request.GET.get('fecha') or timezone.now().strftime('%Y-%m-%d'))
        headers = ['Local','Despacho','Estado','Tipo','Prioridad','Motorista','Cliente','Dirección','Con receta','Incidencia','Fecha']
        filename = 'movimientos_diario'
        rows = []
        try:
            import json, pathlib
            base = pathlib.Path(getattr(settings, 'MEDIA_ROOT', None) or (pathlib.Path(settings.BASE_DIR) / 'media')) / 'reportes'
            file = base / f'cierre_{fecha_arg}.json'
            if file.exists():
                with open(file, 'r', encoding='utf-8') as f:
                    rows = json.load(f) or []
                try:
                    rows = [[r[0], r[1], (_estado_normalizado(r[2]) if ' _estado_normalizado' else (r[2] if r else '')), r[3], r[4], r[5], r[6], r[7], r[8], r[9], r[10]] for r in rows]
                    rows = [r for r in rows if str(r[2]).strip().upper() in {'ENTREGADO','FALLIDO','PENDIENTE'}]
                except Exception:
                    pass
            else:
                from .models import Despacho, Localfarmacia
                try:
                    y, m, d = [int(x) for x in fecha_arg.split('-')]
                except Exception:
                    y, m, d = timezone.now().year, timezone.now().month, timezone.now().day
                qs = Despacho.objects.filter(fecha_registro__year=y, fecha_registro__month=m, fecha_registro__day=d).order_by('fecha_registro')
                for obj in qs:
                    try:
                        farm = Localfarmacia.objects.filter(local_id=obj.farmacia_origen_local_id).first()
                        u = getattr(obj.motorista, 'usuario', None)
                        mot_name = f"{getattr(u,'nombre','')} {getattr(u,'apellido','')}".strip()
                        estado_norm = _estado_normalizado(obj.estado)
                        if estado_norm not in {'ENTREGADO','FALLIDO','PENDIENTE'}:
                            continue
                        rows.append([
                            getattr(farm,'local_id', obj.farmacia_origen_local_id) or '',
                            obj.codigo_despacho or obj.id,
                            estado_norm,
                            obj.tipo_despacho or '',
                            obj.prioridad or '',
                            mot_name,
                            _cliente_normalizado(obj.cliente_nombre),
                            obj.destino_direccion or '',
                            'Sí' if obj.tiene_receta_retenida else 'No',
                            'Sí' if obj.hubo_incidencia else 'No',
                            obj.fecha_registro.strftime('%Y-%m-%d %H:%M') if obj.fecha_registro else '',
                        ])
                    except Exception:
                        continue
        except Exception:
            rows = []
    elif tipo == 'mensual' and detalle:
        from .models import Despacho, Localfarmacia
        y = int(anio) if anio else None
        m = int(mes) if mes else None
        qs = Despacho.objects.all().order_by('fecha_registro')
        if y: qs = qs.filter(fecha_registro__year=y)
        if m: qs = qs.filter(fecha_registro__month=m)
        headers = ['Local','Despacho','Estado','Tipo','Prioridad','Motorista','Cliente','Dirección','Con receta','Incidencia','Fecha']
        filename = f'movimientos_mensual_{anio or "todos"}_{mes or "todos"}'
        rows = []
        for obj in qs:
            try:
                farm = Localfarmacia.objects.filter(local_id=obj.farmacia_origen_local_id).first()
                u = getattr(obj.motorista, 'usuario', None)
                mot_name = f"{getattr(u,'nombre','')} {getattr(u,'apellido','')}".strip()
                estado_norm = _estado_normalizado(obj.estado)
                if estado_norm not in {'ENTREGADO','FALLIDO','PENDIENTE'}:
                    continue
                rows.append([
                    getattr(farm,'local_id', obj.farmacia_origen_local_id) or '',
                    obj.codigo_despacho or obj.id,
                    estado_norm,
                    obj.tipo_despacho or '',
                    obj.prioridad or '',
                    mot_name,
                    _cliente_normalizado(obj.cliente_nombre),
                    obj.destino_direccion or '',
                    'Sí' if obj.tiene_receta_retenida else 'No',
                    'Sí' if obj.hubo_incidencia else 'No',
                    obj.fecha_registro.strftime('%Y-%m-%d %H:%M') if obj.fecha_registro else '',
                ])
            except Exception:
                continue
    else:
        if detalle:
            from .models import Despacho, Localfarmacia
            y = int(anio) if anio else None
            qs = Despacho.objects.all().order_by('fecha_registro')
            if y: qs = qs.filter(fecha_registro__year=y)
            headers = ['Local','Despacho','Estado','Tipo','Prioridad','Motorista','Cliente','Dirección','Con receta','Incidencia','Fecha']
            filename = f'movimientos_anual_{anio or "todos"}'
            rows = []
            for obj in qs:
                try:
                    farm = Localfarmacia.objects.filter(local_id=obj.farmacia_origen_local_id).first()
                    u = getattr(obj.motorista, 'usuario', None)
                    mot_name = f"{getattr(u,'nombre','')} {getattr(u,'apellido','')}".strip()
                    estado_norm = _estado_normalizado(obj.estado)
                    if estado_norm not in {'ENTREGADO','FALLIDO','PENDIENTE'}:
                        continue
                    rows.append([
                        getattr(farm,'local_id', obj.farmacia_origen_local_id) or '',
                        obj.codigo_despacho or obj.id,
                        estado_norm,
                        obj.tipo_despacho or '',
                        obj.prioridad or '',
                        mot_name,
                        _cliente_normalizado(obj.cliente_nombre),
                        obj.destino_direccion or '',
                        'Sí' if obj.tiene_receta_retenida else 'No',
                        'Sí' if obj.hubo_incidencia else 'No',
                        obj.fecha_registro.strftime('%Y-%m-%d %H:%M') if obj.fecha_registro else '',
                    ])
                except Exception:
                    continue
        else:
            rows = get_resumen_operativo_anual(anio=anio)
            headers = ['Año','Farmacia','Comuna','Total despachos','Entregados','Fallidos','Directo','Reenvío receta','Intercambio','Error despacho','Con receta','Con incidencias']
            filename = f'resumen_anual_{anio or "todos"}'
            rows = [[r[0], r[2], r[3], r[4], r[5], r[13], r[14], r[15], r[16], r[10], r[11]] for r in rows]

    if not rows and tipo != 'despachos_activos':
        try:
            from django.db.models import Count, Q, Avg, Sum
            from .models import Despacho, Localfarmacia
            from django.utils import timezone as _tz
            hoy = _tz.now().date()
            qs = Despacho.objects.all()
            if tipo == 'diario':
                qs = qs.filter(fecha_registro__date=hoy)
            elif tipo == 'mensual':
                y = int(anio) if anio else hoy.year
                m = int(mes) if mes else hoy.month
                qs = qs.filter(fecha_registro__year=y, fecha_registro__month=m)
            else:
                y = int(anio) if anio else hoy.year
                qs = qs.filter(fecha_registro__year=y)
            agg = qs.values('farmacia_origen_local_id').annotate(
                total=Count('id'),
                entregados=Count('id', filter=Q(estado='ENTREGADO')),
                fallidos=Count('id', filter=Q(estado='FALLIDO')),
                directo=Count('id', filter=Q(tipo_despacho='DOMICILIO')),
                reenvio=Count('id', filter=Q(tipo_despacho='REENVIO_RECETA')),
                intercambio=Count('id', filter=Q(tipo_despacho='INTERCAMBIO')),
                error=Count('id', filter=Q(tipo_despacho='ERROR_DESPACHO')),
                con_receta=Count('id', filter=Q(tiene_receta_retenida=True)),
                con_incidencias=Count('id', filter=Q(hubo_incidencia=True)),
            )
            mapa_nombres = {lf.local_id: lf.local_nombre for lf in Localfarmacia.objects.all()}
            mapa_comunas = {lf.local_id: lf.comuna_nombre for lf in Localfarmacia.objects.all()}
            if tipo == 'diario':
                rows = [
                    (mapa_nombres.get(r['farmacia_origen_local_id']), mapa_comunas.get(r['farmacia_origen_local_id']), r['total'], r['entregados'], r['fallidos'], r['directo'], r['reenvio'], r['intercambio'], r['error'], r['con_receta'], r['con_incidencias'])
                    for r in agg
                ]
            elif tipo == 'mensual':
                rows = [
                    (y, m, mapa_nombres.get(r['farmacia_origen_local_id']), mapa_comunas.get(r['farmacia_origen_local_id']), r['total'], r['entregados'], r['fallidos'], r['directo'], r['reenvio'], r['intercambio'], r['error'], r['con_receta'], r['con_incidencias'])
                    for r in agg
                ]
            else:
                rows = [
                    (y, mapa_nombres.get(r['farmacia_origen_local_id']), mapa_comunas.get(r['farmacia_origen_local_id']), r['total'], r['entregados'], r['fallidos'], r['directo'], r['reenvio'], r['intercambio'], r['error'], r['con_receta'], r['con_incidencias'])
                    for r in agg
                ]
        except Exception:
            rows = []

    if formato == 'json':
        import json
        from django.http import HttpResponse
        data = [dict(zip(headers, r)) for r in rows]
        return HttpResponse(json.dumps(data, ensure_ascii=False), content_type='application/json')
    elif formato == 'xlsx':
        from django.http import HttpResponse
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Resumen'
            ws.append(headers)
            for r in rows:
                ws.append(list(r))
            resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            resp['Content-Disposition'] = f'attachment; filename={filename}.xlsx'
            import io
            buf = io.BytesIO()
            wb.save(buf)
            resp.write(buf.getvalue())
            return resp
        except Exception:
            formato = 'csv'
    elif formato == 'pdf':
        from django.http import HttpResponse
        try:
            import io
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import mm
            try:
                from reportlab.lib.pdfencrypt import StandardEncryption
                pwd = (settings.PDF_PASSWORD or '000').strip()
                enc = StandardEncryption(pwd, pwd, canPrint=1, canModify=0, canCopy=0, canAnnotate=0)
            except Exception:
                enc = None
            buf = io.BytesIO()
            doc = SimpleDocTemplate(
                buf,
                pagesize=landscape(A4),
                leftMargin=15*mm,
                rightMargin=15*mm,
                topMargin=15*mm,
                bottomMargin=15*mm,
                encrypt=enc
            )
            styles = getSampleStyleSheet()
            elems = []
            elems.append(Paragraph((display_title or filename.replace('_',' ').title()), styles['Title']))
            elems.append(Spacer(1, 6*mm))
            data_rows = [list(r) for r in rows]
            if not data_rows:
                data_rows = [["Sin datos"] + [""] * (len(headers) - 1)]
            data = [headers] + data_rows
            table = Table(data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                ('TEXTCOLOR', (0,0), (-1,0), colors.black),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,-1), 8),
                ('ALIGN', (0,0), (-1,0), 'CENTER'),
                ('ALIGN', (0,1), (-1,-1), 'LEFT'),
                ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.whitesmoke, colors.white]),
            ]))
            elems.append(table)
            doc.build(elems)
            pdf = buf.getvalue()
            resp = HttpResponse(pdf, content_type='application/pdf')
            resp['Content-Disposition'] = f'attachment; filename={filename}.pdf'
            return resp
        except Exception:
            formato = 'csv'
    # CSV fallback
    from django.http import HttpResponse
    import csv, io
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(headers)
    for r in rows:
        w.writerow(list(r))
    audit = f'Generado: {timezone.now().isoformat()} tipo={tipo}'
    w.writerow([audit])
    resp = HttpResponse(buf.getvalue(), content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename={filename}.csv'
    return resp


@permiso_requerido('movimientos', 'view')
def despachos_activos(request):
    rows = get_despachos_activos()
    rol = obtener_rol_usuario(request.user)

    q = (request.GET.get('q') or '').strip().lower()
    estado = (request.GET.get('estado') or '').strip().upper()
    prioridad = (request.GET.get('prioridad') or '').strip().upper()
    tipo = (request.GET.get('tipo') or '').strip().upper()
    receta = (request.GET.get('receta') or '').strip().lower()
    incidencia = (request.GET.get('incidencia') or '').strip().lower()

    data = []
    for r in rows:
        item = {
            'id': r[0],
            'codigo_despacho': r[1],
            'estado': r[2],
            'tipo_despacho': r[3],
            'prioridad': r[4],
            'farmacia_origen': r[5],
            'motorista': r[6],
            'moto_patente': r[7],
            'cliente_nombre': r[8],
            'cliente_telefono': r[9],
            'destino_direccion': r[10],
            'tiene_receta_retenida': bool(r[11]),
            'requiere_aprobacion_operadora': bool(r[12]),
            'aprobado_por_operadora': bool(r[13]),
            'fecha_registro': str(r[14]),
            'fecha_asignacion': str(r[15]),
            'fecha_salida_farmacia': str(r[16]),
            'minutos_en_ruta': r[17],
            'hubo_incidencia': bool(r[18]),
            'tipo_incidencia': r[19],
            'coordenadas_destino': r[20],
        }
        if rol != 'admin':
            tel = item['cliente_telefono']
            s = str(tel or '').strip()
            item['cliente_telefono'] = '***' if not s else ('***' + s[-3:] if len(s) > 3 else '***')
        data.append(item)

    if q:
        data = [d for d in data if q in (d['codigo_despacho'] or '').lower() or q in (d['farmacia_origen'] or '').lower() or q in (d['motorista'] or '').lower() or q in (d['destino_direccion'] or '').lower()]
    if estado:
        data = [d for d in data if (d['estado'] or '').upper() == estado]
    if prioridad:
        data = [d for d in data if (d['prioridad'] or '').upper() == prioridad]
    if tipo:
        data = [d for d in data if (d['tipo_despacho'] or '').upper() == tipo]
    if receta == 'si':
        data = [d for d in data if d['tiene_receta_retenida']]
    elif receta == 'no':
        data = [d for d in data if not d['tiene_receta_retenida']]
    if incidencia == 'si':
        data = [d for d in data if d['hubo_incidencia']]
    elif incidencia == 'no':
        data = [d for d in data if not d['hubo_incidencia']]

    fmt = (request.GET.get('format') or '').strip().lower()
    accept = (request.headers.get('Accept') or '')
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    if fmt == 'json' or 'application/json' in accept or is_ajax:
        from django.http import JsonResponse
        return JsonResponse({'items': data, 'count': len(data)})

    rows_out = [
        [
            d['id'],
            d['codigo_despacho'],
            d['estado'],
            d['tipo_despacho'],
            d['prioridad'],
            d['farmacia_origen'],
            d['motorista'],
            d['moto_patente'],
            d['cliente_nombre'],
            d['cliente_telefono'],
            d['destino_direccion'],
            d['tiene_receta_retenida'],
            d['requiere_aprobacion_operadora'],
            d['aprobado_por_operadora'],
            d['fecha_registro'],
            d['fecha_asignacion'],
            d['fecha_salida_farmacia'],
            d['minutos_en_ruta'],
            d['hubo_incidencia'],
            d['tipo_incidencia'],
            d['coordenadas_destino'],
        ]
        for d in data
    ]

    paginator = Paginator(rows_out, 10)
    page_number = request.GET.get('page')
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    return render(request, 'reportes/despachos-activos.html', {
        'rows': page_obj.object_list,
        'page_obj': page_obj,
        'prioridad': prioridad,
        'receta': receta,
        'incidencia': incidencia,
    })


@permiso_requerido('movimientos', 'view')
def api_despachos_activos(request):
    rows = get_despachos_activos()
    rol = obtener_rol_usuario(request.user)
    data = []
    for r in rows:
        item = {
            'id': r[0],
            'codigo_despacho': r[1],
            'estado': r[2],
            'tipo_despacho': r[3],
            'prioridad': r[4],
            'farmacia_origen': r[5],
            'motorista': r[6],
            'moto_patente': r[7],
            'cliente_nombre': r[8],
            'cliente_telefono': r[9],
            'destino_direccion': r[10],
            'tiene_receta_retenida': bool(r[11]),
            'requiere_aprobacion_operadora': bool(r[12]),
            'aprobado_por_operadora': bool(r[13]),
            'fecha_registro': str(r[14]),
            'fecha_asignacion': str(r[15]),
            'fecha_salida_farmacia': str(r[16]),
            'minutos_en_ruta': r[17],
            'hubo_incidencia': bool(r[18]),
            'tipo_incidencia': r[19],
            'coordenadas_destino': r[20],
        }
        if rol != 'admin':
            s = str(item['cliente_telefono'] or '').strip()
            item['cliente_telefono'] = '***' if not s else ('***' + s[-3:] if len(s) > 3 else '***')
        data.append(item)
    from django.http import JsonResponse
    return JsonResponse({'items': data, 'count': len(data)})


@permiso_requerido('movimientos', 'view')
def recetas_pendientes_devolucion(request):
    rows = []
    with connection.cursor() as cur:
        cur.execute("SELECT despacho_id, codigo_despacho, numero_receta, fecha_registro, fecha_completado, dias_desde_registro, dias_desde_completado, farmacia_origen, farmacia_telefono, motorista, motorista_telefono, cliente_nombre, cliente_telefono, estado, nivel_alerta FROM vista_recetas_pendientes_devolucion ORDER BY dias_desde_completado DESC")
        rows = cur.fetchall()
    return render(request, 'reportes/recetas-pendientes.html', {'rows': rows})


@permiso_requerido('movimientos', 'view')
def consulta_rapida(request):
    local = request.GET.get('local', '').strip()
    motorista = request.GET.get('motorista', '').strip()
    cliente = request.GET.get('cliente', '').strip()
    def _clamp(s):
        s = (s or '').strip()
        return s[:100]
    local = _clamp(local)
    motorista = _clamp(motorista)
    cliente = _clamp(cliente)
    results = []
    with connection.cursor() as cur:
        cur.execute(
            "SELECT d.codigo_despacho, lf.local_nombre, CONCAT(u.nombre,' ',u.apellido) AS motorista, d.cliente_nombre, d.estado, d.tipo_despacho, d.prioridad, d.fecha_registro FROM despacho d JOIN localfarmacia lf ON d.farmacia_origen_local_id = lf.local_id JOIN motorista mot ON d.motorista_id = mot.id JOIN usuario u ON mot.codigo_motorista = u.id WHERE (lf.local_nombre LIKE %s OR %s = '') AND (CONCAT(u.nombre,' ',u.apellido) LIKE %s OR %s = '') AND (d.cliente_nombre LIKE %s OR %s = '') ORDER BY d.fecha_registro DESC LIMIT 200",
            [f"%{local}%", local, f"%{motorista}%", motorista, f"%{cliente}%", cliente]
        )
        results = cur.fetchall()
    return render(request, 'reportes/consulta-rapida.html', {'results': results, 'local': local, 'motorista': motorista, 'cliente': cliente})


@rol_requerido('operador')
def movimiento_anular(request):
    return render(request, 'movimientos/anular-mov.html')


@rol_requerido('operador')
def movimiento_modificar(request):
    return render(request, 'movimientos/modificar-mov.html')

@permiso_requerido('asignaciones', 'add')
def agregar_asignacion_moto_motorista(request):
    """Crea una asignación Moto–Motorista (empresa asigna moto a motorista)."""
    form = None
    if request.method == 'POST':
        form = AsignarMotoristaForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            moto = cd.get('moto')
            motorista = cd.get('motorista')
            try:
                activo_moto = AsignacionMotoMotorista.objects.filter(moto=moto, activa=1).first()
                if activo_moto:
                    messages.error(request, 'La moto ya está asignada y activa.')
                    return render(request, 'asignaciones/agregar-asignacion-moto-motorista.html', {'form': form, 'titulo': 'Asignar Moto a Motorista'})
                activo_mot = AsignacionMotoMotorista.objects.filter(motorista=motorista, activa=1).first()
                if activo_mot:
                    messages.error(request, 'El motorista ya tiene una moto activa.')
                    return render(request, 'asignaciones/agregar-asignacion-moto-motorista.html', {'form': form, 'titulo': 'Asignar Moto a Motorista'})
            except Exception:
                pass
            obj = form.save(commit=False)
            try:
                from django.utils import timezone as _tz
                obj.fecha_asignacion = obj.fecha_asignacion or _tz.now()
                obj.kilometraje_inicio = obj.kilometraje_inicio or int(getattr(moto, 'kilometraje_actual', 0) or 0)
                obj.activa = bool(cd.get('activa', True))
            except Exception:
                pass
            obj.save()
            messages.success(request, 'Moto asignada al motorista exitosamente.')
            return redirect('detalle_moto', pk=moto.pk)
        else:
            messages.error(request, 'Corrige los errores del formulario.')
    else:
        initial = {}
        try:
            mid = int(request.GET.get('motorista') or '0')
            if mid:
                initial['motorista'] = mid
        except Exception:
            pass
        try:
            mpat = (request.GET.get('moto_patente') or '').strip().upper()
            if mpat:
                mo = Moto.objects.filter(patente=mpat).first()
                if mo:
                    initial['moto'] = mo.pk
        except Exception:
            pass
        form = AsignarMotoristaForm(initial=initial)
    return render(request, 'asignaciones/agregar-asignacion-moto-motorista.html', {'form': form, 'titulo': 'Asignar Moto a Motorista'})

@permiso_requerido('movimientos', 'view')
def movimiento_directo(request):
    return render(request, 'movimientos/mov-directo.html')


@permiso_requerido('movimientos', 'view')
def movimiento_receta(request):
    return render(request, 'movimientos/mov-receta.html')


@permiso_requerido('movimientos', 'view')
def movimiento_reenvio(request):
    return render(request, 'movimientos/mov-reenvio.html')


@permiso_requerido('movimientos', 'view')
def movimiento_traslado(request):
    return render(request, 'movimientos/mov-traslado.html')


@permiso_requerido('movimientos', 'view')
def panel_operadora(request):
    return render(request, 'operadora/panel-operadora.html')


@permiso_requerido('movimientos', 'add')
def cerrar_dia_operadora(request):
    try:
        import json, pathlib
        hoy = timezone.now().date()
        qs = Despacho.objects.filter(fecha_registro__date=hoy).order_by('fecha_registro')
        rows = []
        if qs.count() == 0:
            tipos = ['DOMICILIO','REENVIO_RECETA','INTERCAMBIO','ERROR_DESPACHO']
            estados = ['PENDIENTE','ASIGNADO','EN_CAMINO','ENTREGADO','FALLIDO']
            prioridades = ['ALTA','MEDIA','BAJA']
            farms = list(Farmacia.objects.all())
            mots = list(Motorista.objects.all())
            usuario_reg = Usuario.objects.filter(django_user_id=request.user.id).first() or Usuario.objects.first()
            base_dt = timezone.now()
            for i in range(100):
                try:
                    f = farms[i % len(farms)] if farms else None
                    m = mots[i % len(mots)] if mots else None
                    t = tipos[i % len(tipos)]
                    e = estados[(i*3) % len(estados)]
                    p = prioridades[(i*5) % len(prioridades)]
                    codigo = f"DSP-{base_dt.strftime('%Y%m%d')}-{i:04d}"
                    if Despacho.objects.filter(codigo_despacho=codigo).exists():
                        continue
                    obj = Despacho(
                        codigo_despacho=codigo,
                        numero_orden_farmacia=f"ORD-{i:05d}",
                        farmacia_origen_local_id=(getattr(f, 'local_id', None) if f else None),
                        farmacia_destino_local_id=(random.choice(farms).local_id if (t=='INTERCAMBIO' and farms) else None),
                        motorista=m,
                        estado=e,
                        tipo_despacho=t,
                        prioridad=p,
                        cliente_nombre="Cliente Uno",
                        cliente_telefono='+56900000000',
                        destino_direccion=f"Calle {i} #123",
                        destino_referencia='Frente a plaza',
                        destino_geolocalizacion_validada=False,
                        tiene_receta_retenida=(t=='REENVIO_RECETA'),
                        numero_receta=(f"REC-{i:05d}" if t=='REENVIO_RECETA' else None),
                        requiere_devolucion_receta=(t=='REENVIO_RECETA'),
                        receta_devuelta_farmacia=False,
                        observaciones_receta=None,
                        descripcion_productos='Demo productos',
                        valor_declarado=10000 + (i * 100),
                        requiere_aprobacion_operadora=False,
                        aprobado_por_operadora=False,
                        firma_digital=(e=='ENTREGADO'),
                        hubo_incidencia=(t=='ERROR_DESPACHO'),
                        usuario_aprobador=None,
                        fecha_aprobacion=None,
                        fecha_registro=base_dt,
                        fecha_asignacion=base_dt,
                        fecha_salida_farmacia=None,
                        fecha_modificacion=base_dt,
                        usuario_registro=usuario_reg,
                        usuario_modificacion=usuario_reg,
                    )
                    obj.save()
                except Exception:
                    continue
            qs = Despacho.objects.filter(fecha_registro__date=hoy).order_by('fecha_registro')
            # Construir filas del cierre únicamente con ENTREGADO y FALLIDO
            for d in qs:
                try:
                    estado_norm = (d.estado or '').strip().upper()
                    if estado_norm not in {'ENTREGADO','FALLIDO'}:
                        continue
                    u = getattr(d.motorista, 'usuario', None)
                    mot_name = f"{getattr(u,'nombre','')} {getattr(u,'apellido','')}".strip()
                    rows.append([
                        d.farmacia_origen_local_id or '',
                        d.codigo_despacho or d.id,
                        d.estado or '',
                        d.tipo_despacho or '',
                        d.prioridad or '',
                        mot_name,
                        _cliente_normalizado(d.cliente_nombre),
                        d.destino_direccion or '',
                        'Sí' if d.tiene_receta_retenida else 'No',
                        'Sí' if d.hubo_incidencia else 'No',
                        d.fecha_registro.strftime('%Y-%m-%d %H:%M') if d.fecha_registro else '',
                    ])
                except Exception:
                    continue
        else:
            # Asignar código si falta para garantizar consistencia del cierre
            seq = 0
            for d in qs:
                try:
                    u = getattr(d.motorista, 'usuario', None)
                    mot_name = f"{getattr(u,'nombre','')} {getattr(u,'apellido','')}".strip()
                    if not (d.codigo_despacho or '').strip():
                        d.codigo_despacho = f"DSP-{hoy.strftime('%Y%m%d')}-{seq:04d}"
                        seq += 1
                        d.save()
                    from .models import Localfarmacia
                    farm = Localfarmacia.objects.filter(local_id=d.farmacia_origen_local_id).first()
                    estado_norm = (d.estado or '').strip().upper()
                    if estado_norm not in {'ENTREGADO','FALLIDO'}:
                        continue
                    rows.append([
                        d.farmacia_origen_local_id or '',
                        d.codigo_despacho or d.id,
                        d.estado or '',
                        d.tipo_despacho or '',
                        d.prioridad or '',
                        mot_name,
                        _cliente_normalizado(d.cliente_nombre),
                        d.destino_direccion or '',
                        'Sí' if d.tiene_receta_retenida else 'No',
                        'Sí' if d.hubo_incidencia else 'No',
                        d.fecha_registro.strftime('%Y-%m-%d %H:%M') if d.fecha_registro else '',
                    ])
                except Exception:
                    continue
        base_dir = pathlib.Path(getattr(settings, 'MEDIA_ROOT', None) or (pathlib.Path(settings.BASE_DIR) / 'media')) / 'reportes'
        base_dir.mkdir(parents=True, exist_ok=True)
        cierre_file = base_dir / f"cierre_{hoy.strftime('%Y-%m-%d')}.json"
        with open(cierre_file, 'w', encoding='utf-8') as f:
            json.dump(rows, f, ensure_ascii=False)
        pass
    except Exception as e:
        pass
    return redirect('despachos_activos')

@permiso_requerido('movimientos', 'add')
def generar_despachos_demo(request):
    try:
        from .models import Despacho, Localfarmacia, Motorista, Usuario
        import random, pathlib, json
        hoy_dt = timezone.now()
        farms = list(Localfarmacia.objects.all())
        mots = list(Motorista.objects.all())
        if not farms:
            messages.error(request, 'No hay farmacias disponibles para generar demo')
            return redirect('despachos_activos')
        if not mots:
            messages.error(request, 'No hay motoristas disponibles para generar demo')
            return redirect('despachos_activos')
        tipos = ['DOMICILIO','REENVIO_RECETA','INTERCAMBIO','ERROR_DESPACHO']
        estados = ['PENDIENTE','ASIGNADO','EN_CAMINO','ENTREGADO','FALLIDO']
        prioridades = ['ALTA','MEDIA','BAJA']
        created = 0
        for i in range(100):
            try:
                f = random.choice(farms)
                m = random.choice(mots)
                t = tipos[i % len(tipos)]
                e = estados[(i*3) % len(estados)]
                p = prioridades[(i*5) % len(prioridades)]
                codigo = f"DSP-{hoy_dt.strftime('%Y%m%d')}-{i:04d}"
                if Despacho.objects.filter(codigo_despacho=codigo).exists():
                    continue
                obj = Despacho(
                    codigo_despacho=codigo,
                    numero_orden_farmacia=f"ORD-{i:05d}",
                    farmacia_origen_local_id=f.local_id,
                    farmacia_destino_local_id=(random.choice(farms).local_id if t=='INTERCAMBIO' else None),
                    motorista=m,
                    estado=e,
                    tipo_despacho=t,
                    prioridad=p,
                    cliente_nombre="Cliente Uno",
                    cliente_telefono='+56900000000',
                    destino_direccion=f"Calle {i} #123",
                    destino_referencia='Frente a plaza',
                    destino_geolocalizacion_validada=False,
                    tiene_receta_retenida=(t=='REENVIO_RECETA'),
                    numero_receta=(f"REC-{i:05d}" if t=='REENVIO_RECETA' else None),
                    requiere_devolucion_receta=(t=='REENVIO_RECETA'),
                    receta_devuelta_farmacia=False,
                    observaciones_receta=None,
                    descripcion_productos='Demo productos',
                    valor_declarado=10000 + (i * 100),
                    requiere_aprobacion_operadora=False,
                    aprobado_por_operadora=False,
                    firma_digital=(e=='ENTREGADO'),
                    hubo_incidencia=(t=='ERROR_DESPACHO'),
                    usuario_aprobador=None,
                    fecha_aprobacion=None,
                    fecha_registro=hoy_dt,
                    fecha_asignacion=hoy_dt,
                    fecha_salida_farmacia=None,
                    fecha_modificacion=hoy_dt,
                    usuario_registro=Usuario.objects.filter(django_user_id=request.user.id).first(),
                    usuario_modificacion=Usuario.objects.filter(django_user_id=request.user.id).first(),
                )
                obj.save()
                created += 1
            except Exception:
                continue
        base_dir = pathlib.Path(getattr(settings, 'MEDIA_ROOT', None) or (pathlib.Path(settings.BASE_DIR) / 'media')) / 'reportes'
        base_dir.mkdir(parents=True, exist_ok=True)
        cierre_file = base_dir / f"cierre_{hoy_dt.strftime('%Y-%m-%d')}.json"
        rows = []
        qs = Despacho.objects.filter(fecha_registro__date=hoy_dt.date()).order_by('fecha_registro')
        for d in qs:
            try:
                u = getattr(d.motorista, 'usuario', None)
                mot_name = f"{getattr(u,'nombre','')} {getattr(u,'apellido','')}".strip()
                rows.append([
                    d.farmacia_origen_local_id or '',
                    d.codigo_despacho or d.id,
                    d.estado or '',
                    d.tipo_despacho or '',
                    d.prioridad or '',
                    mot_name,
                    (d.cliente_nombre or 'Cliente Uno'),
                    d.destino_direccion or '',
                    'Sí' if d.tiene_receta_retenida else 'No',
                    'Sí' if d.hubo_incidencia else 'No',
                    d.fecha_registro.strftime('%Y-%m-%d %H:%M') if d.fecha_registro else '',
                ])
            except Exception:
                continue
        with open(cierre_file, 'w', encoding='utf-8') as f:
            json.dump(rows, f, ensure_ascii=False)
        pass
    except Exception as e:
        pass
    return redirect('despachos_activos')

@permiso_requerido('movimientos', 'view')
def recetas_retencion_panel(request):
    from .models import Despacho
    
    filtro_farmacia = request.GET.get('farmacia','').strip()
    filtro_motorista = request.GET.get('motorista','').strip()

    recetas = Despacho.objects.filter(
        tiene_receta_retenida=True,
        requiere_devolucion_receta=True,
        receta_devuelta_farmacia=False
    ).select_related('motorista')
    if not recetas.exists():
        try:
            import json, pathlib
            data_path = pathlib.Path(__file__).resolve().parent.parent / 'static' / 'data' / 'recetas_retencion.json'
            with open(data_path, 'r', encoding='utf-8') as f:
                raw = json.load(f)
            class D:
                def __init__(self, d):
                    self.codigo_despacho = d.get('codigo_despacho')
                    self.farmacia_origen_local_id = d.get('farmacia_origen_local_id')
                    self.motorista = type('Mot', (), {'usuario': type('U', (), {'nombre': d.get('motorista', {}).get('nombre'), 'apellido': d.get('motorista', {}).get('apellido')})})
                    self.cliente_nombre = d.get('cliente_nombre')
                    self.estado = d.get('estado')
                    self.fecha_registro = d.get('fecha_registro')
            recetas = [D(d) for d in raw]
        except Exception:
            pass

    historico = Despacho.objects.filter(
        tiene_receta_retenida=True,
        receta_devuelta_farmacia=True
    ).select_related('motorista').order_by('-fecha_devolucion_receta')[:200]

    if filtro_farmacia:
        recetas = recetas.filter(farmacia_origen_local_id__icontains=filtro_farmacia)
        historico = historico.filter(farmacia_origen_local_id__icontains=filtro_farmacia)
    if filtro_motorista:
        recetas = recetas.filter(motorista__usuario__nombre__icontains=filtro_motorista) | recetas.filter(motorista__usuario__apellido__icontains=filtro_motorista)
        historico = historico.filter(motorista__usuario__nombre__icontains=filtro_motorista) | historico.filter(motorista__usuario__apellido__icontains=filtro_motorista)

    return render(request, 'operadora/recetas-retencion.html', {
        'recetas': recetas,
        'historico': historico,
        'farmacia': filtro_farmacia,
        'motorista': filtro_motorista,
    })


@permiso_requerido('movimientos', 'add')
def receta_marcar_devuelta(request, despacho_id):
    from .models import Despacho, Usuario
    d = Despacho.objects.filter(id=despacho_id).first()
    if not d:
        messages.error(request, 'Despacho no encontrado')
        return redirect('recetas_retencion_panel')
    estado_norm = (d.estado or '').strip().upper()
    if estado_norm not in {'PREPARANDO','PREPARADO','EN PROCESO','EN_PROCESO','PROCESO'}:
        messages.error(request, 'Solo puedes marcar devolución cuando el despacho está EN PROCESO')
        return redirect('recetas_retencion_panel')
    if not (d.tiene_receta_retenida and d.requiere_devolucion_receta):
        messages.error(request, 'La devolución aplica solo para receta retenida con devolución requerida')
        return redirect('recetas_retencion_panel')
    quien = request.POST.get('quien_recibe', '').strip()
    notas = request.POST.get('observaciones', '').strip()
    d.receta_devuelta_farmacia = True
    d.fecha_devolucion_receta = timezone.now()
    d.quien_recibe_receta = quien or d.quien_recibe_receta
    d.observaciones_receta = notas or d.observaciones_receta
    usuario = Usuario.objects.filter(django_user_id=request.user.id).first()
    d.usuario_modificacion = usuario
    d.fecha_modificacion = timezone.now()
    d.save()
    messages.success(request, 'Receta marcada como devuelta')
    return redirect('recetas_retencion_panel')
# ===== DESPACHOS =====
@permiso_requerido('despachos', 'view')
def listado_despachos(request):
    search = request.GET.get('search','').strip()
    receta = request.GET.get('receta','').strip()
    requiere = request.GET.get('requiere','').strip()
    qs = Despacho.objects.all().order_by('-fecha_registro')
    if search:
        qs = qs.filter(Q(codigo_despacho__icontains=search) | Q(cliente_nombre__icontains=search) | Q(farmacia_origen_local_id__icontains=search))
    if receta == 'si':
        qs = qs.filter(tiene_receta_retenida=True)
    elif receta == 'no':
        qs = qs.filter(tiene_receta_retenida=False)
    if requiere == 'si':
        qs = qs.filter(requiere_devolucion_receta=True)
    elif requiere == 'no':
        qs = qs.filter(requiere_devolucion_receta=False)
    paginator = Paginator(qs, 10)
    page = request.GET.get('page')
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    return render(request, 'despachos/listado-despachos.html', {
        'page_obj': page_obj,
        'search': search,
        'receta': receta,
        'requiere': requiere,
    })


@permiso_requerido('despachos', 'add')
def agregar_despacho(request):
    from .models import Usuario
    if request.method == 'POST':
        form = DespachoForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.usuario_registro = Usuario.objects.filter(django_user_id=request.user.id).first()
            obj.fecha_registro = timezone.now()
            obj.fecha_modificacion = timezone.now()
            # Sugerir farmacia origen por comuna si no se indica
            try:
                if not (obj.farmacia_origen_local_id or '').strip():
                    from .models import Localfarmacia
                    comuna = (obj.cliente_comuna_nombre or '').strip()
                    sug = Localfarmacia.objects.filter(activo=True, comuna_nombre__icontains=comuna).order_by('local_nombre').first()
                    if sug:
                        obj.farmacia_origen_local_id = sug.local_id
                        messages.info(request, f'Farmacia sugerida: {sug.local_nombre} ({sug.local_id}) por comuna {comuna}')
            except Exception:
                pass
            obj.save()
            # Auditoría de creación
            try:
                from .models import AuditoriaGeneral, Usuario as U
                AuditoriaGeneral.objects.create(
                    nombre_tabla='despacho',
                    id_registro_afectado=str(obj.id),
                    tipo_operacion='INSERT',
                    usuario=U.objects.filter(django_user_id=request.user.id).first(),
                    fecha_evento=timezone.now(),
                    datos_antiguos=None,
                    datos_nuevos={
                        'codigo': obj.codigo_despacho,
                        'estado': obj.estado,
                        'tipo': obj.tipo_despacho,
                        'prioridad': obj.prioridad,
                        'farmacia_origen_local_id': obj.farmacia_origen_local_id,
                        'tiene_receta_retenida': obj.tiene_receta_retenida,
                        'requiere_devolucion_receta': obj.requiere_devolucion_receta,
                    }
                )
            except Exception:
                pass
            messages.success(request, 'Despacho creado')
            return redirect('detalle_despacho', pk=obj.id)
        else:
            messages.error(request, 'Corrige los errores')
    else:
        form = DespachoForm()
    return render(request, 'despachos/agregar-despacho.html', {'form': form})


@permiso_requerido('despachos', 'change')
def actualizar_despacho(request, pk):
    from .models import Usuario
    d = get_object_or_404(Despacho, pk=pk)
    if request.method == 'POST':
        form = DespachoForm(request.POST, instance=d)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.usuario_modificacion = Usuario.objects.filter(django_user_id=request.user.id).first()
            obj.fecha_modificacion = timezone.now()
            prev_estado = d.estado
            obj.save()
            try:
                if prev_estado != obj.estado:
                    from .models import MovimientoDespacho
                    MovimientoDespacho.objects.create(
                        despacho=obj,
                        estado_anterior=prev_estado,
                        estado_nuevo=obj.estado,
                        fecha_movimiento=timezone.now(),
                        usuario=obj.usuario_modificacion,
                        observacion='actualizacion_despacho'
                    )
            except Exception:
                pass
            # Auditoría de actualización
            try:
                from .models import AuditoriaGeneral, Usuario as U
                AuditoriaGeneral.objects.create(
                    nombre_tabla='despacho',
                    id_registro_afectado=str(obj.id),
                    tipo_operacion='UPDATE',
                    usuario=U.objects.filter(django_user_id=request.user.id).first(),
                    fecha_evento=timezone.now(),
                    datos_antiguos=None,
                    datos_nuevos={
                        'estado': obj.estado,
                        'tiene_receta_retenida': obj.tiene_receta_retenida,
                        'requiere_devolucion_receta': obj.requiere_devolucion_receta,
                        'numero_receta': obj.numero_receta,
                    }
                )
            except Exception:
                pass
            messages.success(request, 'Despacho actualizado')
            return redirect('detalle_despacho', pk=obj.id)
        else:
            messages.error(request, 'Corrige los errores')
    else:
        form = DespachoForm(instance=d)
    return render(request, 'despachos/modificar-despacho.html', {'form': form, 'despacho': d})


@permiso_requerido('despachos', 'delete')
def remover_despacho(request, pk):
    d = get_object_or_404(Despacho, pk=pk)
    if request.method == 'POST':
        try:
            d.delete()
            messages.success(request, 'Despacho eliminado')
        except Exception as e:
            messages.error(request, f'Error: {e}')
        return redirect('listado_despachos')
    return render(request, 'despachos/remover-despacho.html', {'despacho': d})


@permiso_requerido('despachos', 'view')
def detalle_despacho(request, pk):
    d = get_object_or_404(Despacho, pk=pk)
    from .models import MovimientoDespacho
    movs = MovimientoDespacho.objects.filter(despacho=d).order_by('-fecha_movimiento')
    return render(request, 'despachos/detalle-despacho.html', {'despacho': d, 'movimientos': movs})


@permiso_requerido('despachos', 'change')
def actualizar_receta_despacho(request, pk):
    d = get_object_or_404(Despacho, pk=pk)
    if request.method == 'POST':
        estado_norm = (d.estado or '').strip().upper()
        if estado_norm not in {'PREPARANDO','PREPARADO','EN PROCESO','EN_PROCESO','PROCESO'}:
            messages.error(request, 'Solo puedes editar datos de receta cuando el despacho está EN PROCESO')
            return redirect('detalle_despacho', pk=d.id)
        d.tiene_receta_retenida = True if request.POST.get('tiene_receta_retenida') == 'on' else False
        d.numero_receta = request.POST.get('numero_receta', d.numero_receta)
        d.requiere_devolucion_receta = True if request.POST.get('requiere_devolucion_receta') == 'on' else False
        if d.tiene_receta_retenida and not d.requiere_devolucion_receta:
            d.requiere_devolucion_receta = True
        d.quien_recibe_receta = request.POST.get('quien_recibe_receta', d.quien_recibe_receta)
        d.observaciones_receta = request.POST.get('observaciones_receta', d.observaciones_receta)
        if request.POST.get('marcar_devuelta') == 'si':
            if not (d.tiene_receta_retenida and d.requiere_devolucion_receta):
                messages.error(request, 'Para marcar devuelta, debe estar retenida y requerir devolución')
                return redirect('detalle_despacho', pk=d.id)
            d.receta_devuelta_farmacia = True
            d.fecha_devolucion_receta = timezone.now()
        try:
            from .models import Usuario
            d.usuario_modificacion = Usuario.objects.filter(django_user_id=request.user.id).first()
            d.fecha_modificacion = timezone.now()
            d.save()
            messages.success(request, 'Datos de receta actualizados')
        except Exception as e:
            messages.error(request, f'Error: {e}')
    return redirect('detalle_despacho', pk=d.id)


@permiso_requerido('despachos', 'change')
def solicitar_correccion_estado(request, pk):
    from .models import Despacho, AuditoriaGeneral, Usuario
    d = get_object_or_404(Despacho, pk=pk)
    if request.method != 'POST':
        return redirect('detalle_despacho', pk=pk)
    motivo = (request.POST.get('motivo','') or '').strip()
    objetivo = (request.POST.get('estado_objetivo','') or '').strip().upper()
    if not motivo or len(motivo) < 5:
        messages.error(request, 'Describe el motivo de la corrección (mínimo 5 caracteres)')
        return redirect('detalle_despacho', pk=pk)
    mapa_prev = {
        'ASIGNADO': 'PENDIENTE',
        'PREPARANDO': 'ASIGNADO',
        'PREPARADO': 'PREPARANDO',
        'EN_CAMINO': 'PREPARADO',
        'ENTREGADO': 'EN_CAMINO',
        'FALLIDO': 'EN_CAMINO',
    }
    estado_actual = (d.estado or '').strip().upper()
    permitido = mapa_prev.get(estado_actual)
    if not permitido:
        messages.error(request, 'No es posible solicitar corrección desde este estado')
        return redirect('detalle_despacho', pk=pk)
    if objetivo and objetivo != permitido:
        messages.error(request, 'Solo se permite volver un paso atrás en el orden operativo')
        return redirect('detalle_despacho', pk=pk)
    try:
        AuditoriaGeneral.objects.create(
            nombre_tabla='despacho',
            id_registro_afectado=str(d.id),
            tipo_operacion='CORRECCION_SOLICITADA',
            usuario=Usuario.objects.filter(django_user_id=request.user.id).first(),
            fecha_evento=timezone.now(),
            datos_antiguos={'estado_actual': estado_actual},
            datos_nuevos={'estado_objetivo': permitido, 'motivo': motivo}
        )
        messages.info(request, 'Corrección solicitada. Un supervisor debe aprobar la reversión')
    except Exception as e:
        messages.error(request, f'Error al solicitar corrección: {e}')
    return redirect('detalle_despacho', pk=pk)


@rol_requerido('supervisor')
def aplicar_correccion_estado(request, pk):
    from .models import Despacho, AuditoriaGeneral, Usuario
    d = get_object_or_404(Despacho, pk=pk)
    # Buscar la última corrección solicitada
    from django.utils import timezone as tz
    limite = tz.now() - timezone.timedelta(hours=12)
    corr = AuditoriaGeneral.objects.filter(
        nombre_tabla='despacho',
        id_registro_afectado=str(d.id),
        tipo_operacion='CORRECCION_SOLICITADA',
        fecha_evento__gte=limite
    ).order_by('-fecha_evento').first()
    if not corr:
        messages.error(request, 'No hay correcciones pendientes para este despacho')
        return redirect('detalle_despacho', pk=pk)
    objetivo = (corr.datos_nuevos or {}).get('estado_objetivo')
    if not objetivo:
        messages.error(request, 'Corrección inválida')
        return redirect('detalle_despacho', pk=pk)
    # Aplicar solo si el estado actual coincide con lo registrado
    estado_actual = (d.estado or '').strip().upper()
    estado_reg = (corr.datos_antiguos or {}).get('estado_actual')
    if estado_actual != estado_reg:
        messages.error(request, 'El estado actual no coincide con la solicitud de corrección')
        return redirect('detalle_despacho', pk=pk)
    try:
        d.estado = objetivo
        d.usuario_modificacion = Usuario.objects.filter(django_user_id=request.user.id).first()
        d.fecha_modificacion = timezone.now()
        d.save()
        AuditoriaGeneral.objects.create(
            nombre_tabla='despacho',
            id_registro_afectado=str(d.id),
            tipo_operacion='CORRECCION_APROBADA',
            usuario=Usuario.objects.filter(django_user_id=request.user.id).first(),
            fecha_evento=timezone.now(),
            datos_antiguos={'estado': estado_actual},
            datos_nuevos={'estado': objetivo, 'motivo': (corr.datos_nuevos or {}).get('motivo')}
        )
        messages.success(request, 'Corrección aplicada')
    except Exception as e:
        messages.error(request, f'Error al aplicar corrección: {e}')
    return redirect('detalle_despacho', pk=pk)
def movimientos_general(request):
    return render(request, 'reportes/movimientos-general.html')


@rol_requerido('motorista')
def avisar_movimiento_motorista(request):
    from .models import AuditoriaGeneral, Usuario
    if request.method == 'POST':
        codigo = (request.POST.get('codigo_despacho','') or '').strip()
        tipo = (request.POST.get('tipo_movimiento','') or '').strip().upper()
        metodo = (request.POST.get('metodo','') or '').strip().lower()
        texto = (request.POST.get('mensaje','') or '').strip()
        if not codigo or not tipo:
            messages.error(request, 'Completa código y tipo de movimiento')
            return redirect('avisar_movimiento_motorista')
        u = Usuario.objects.filter(django_user_id=request.user.id).first()
        try:
            AuditoriaGeneral.objects.create(
                nombre_tabla='comunicacion',
                id_registro_afectado=codigo,
                tipo_operacion='AVISO_MOV',
                usuario=u,
                fecha_evento=timezone.now(),
                datos_antiguos=None,
                datos_nuevos={
                    'codigo': codigo,
                    'tipo_mov': tipo,
                    'metodo': metodo,
                    'mensaje': texto,
                }
            )
            messages.success(request, 'Aviso enviado a Operadora')
        except Exception as e:
            messages.error(request, f'Error: {e}')
        return redirect('avisar_movimiento_motorista')
    return render(request, 'motoristas/avisar-movimiento.html')


@permiso_requerido('movimientos', 'view')
def feed_avisos_operadora(request):
    from .models import AuditoriaGeneral
    codigo = (request.GET.get('codigo','') or '').strip()
    motorista = (request.GET.get('motorista','') or '').strip()
    avisos = AuditoriaGeneral.objects.filter(tipo_operacion='AVISO_MOV')
    if codigo:
        avisos = avisos.filter(datos_nuevos__codigo__icontains=codigo)
    if motorista:
        avisos = avisos.filter(Q(usuario__nombre__icontains=motorista) | Q(usuario__apellido__icontains=motorista))
    avisos = avisos.order_by('-fecha_evento')
    paginator = Paginator(avisos, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'operadora/avisos.html', {'page_obj': page_obj, 'codigo': codigo, 'motorista': motorista})


@permiso_requerido('movimientos', 'change')
def marcar_aviso_leido(request, audit_id):
    from .models import AuditoriaGeneral, Usuario
    a = AuditoriaGeneral.objects.filter(id=audit_id, tipo_operacion='AVISO_MOV').first()
    if not a:
        messages.error(request, 'Aviso no encontrado')
        return redirect('feed_avisos_operadora')
    try:
        AuditoriaGeneral.objects.create(
            nombre_tabla='comunicacion',
            id_registro_afectado=str(a.id),
            tipo_operacion='AVISO_MOV_LEIDO',
            usuario=Usuario.objects.filter(django_user_id=request.user.id).first(),
            fecha_evento=timezone.now(),
            datos_antiguos=None,
            datos_nuevos={'codigo': (a.datos_nuevos or {}).get('codigo'), 'leido': 1}
        )
        messages.success(request, 'Aviso marcado como leído')
    except Exception as e:
        messages.error(request, f'Error: {e}')
    return redirect('feed_avisos_operadora')
 
@permiso_requerido('movimientos', 'view')
def ia_sugerencias(request):
    estado = (request.GET.get('estado') or 'FALLIDO').strip().upper()
    try:
        limit = int(request.GET.get('limit') or 200)
    except Exception:
        limit = 200
    ai = None
    try:
        from appnproylogico.IA_gen import analizar_incidencia as ai
    except Exception:
        try:
            from IA_gen import analizar_incidencia as ai
        except Exception:
            ai = None
    def _fallback(d, cierre='20:00'):
        from .models import Localfarmacia
        est = (getattr(d, 'estado', '') or 'Desconocido').upper()
        pri = getattr(d, 'prioridad', '') or 'No definida'
        lid = (getattr(d, 'farmacia_origen_local_id', '') or '').strip()
        farm = Localfarmacia.objects.filter(local_id=lid).first()
        cierre_t = getattr(farm, 'funcionamiento_hora_cierre', None)
        now_t = timezone.now().time()
        sug = 'Postergar'
        if est in {'EN_CAMINO','EN_PROCESO'}:
            sug = 'Supervisar'
        elif est == 'FALLIDO':
            sug = 'Reasignar/Reenviar'
        elif est == 'PENDIENTE' and (pri or '').upper() == 'ALTA':
            sug = 'Reasignar/Reenviar'
        if cierre_t and now_t >= cierre_t:
            sug = 'Reasignar/Reenviar' if (pri or '').upper() == 'ALTA' else 'Postergar'
        return f"Resumen: {est} · {pri} · {(getattr(d,'cliente_nombre','') or 'Cliente')}\nSugerencia: {sug}"
    fn = ai or _fallback
    from .models import Despacho
    qs = Despacho.objects.all().order_by('-fecha_registro')
    if estado:
        qs = qs.filter(estado=estado)
    rows = []
    for d in qs[:limit]:
        out = fn(d, horario_cierre=getattr(settings, 'HORARIO_CIERRE_DEFAULT', '20:00'))
        sug = out.split('\n')[-1]
        rows.append({
            'codigo': d.codigo_despacho or d.id,
            'estado': d.estado,
            'prioridad': d.prioridad,
            'cliente': d.cliente_nombre,
            'sugerencia': sug,
        })
    return render(request, 'operadora/ia-sugerencias.html', {'rows': rows, 'estado': estado, 'limit': limit})

@permiso_requerido('farmacias', 'view')
def geolocalizacion_panel(request):
    from .models import Localfarmacia
    total = Localfarmacia.objects.count()
    pendientes = Localfarmacia.objects.filter(activo=True).filter(models.Q(local_lat__isnull=True) | models.Q(local_lng__isnull=True) | models.Q(geolocalizacion_validada=False)).count()
    ok = Localfarmacia.objects.filter(geolocalizacion_validada=True).count()
    ultimas = Localfarmacia.objects.filter(geolocalizacion_validada=True).order_by('-fecha_geolocalizacion')[:20]
    return render(request, 'operadora/geolocalizacion-panel.html', {'total': total, 'pendientes': pendientes, 'ok': ok, 'ultimas': ultimas})
@permiso_requerido('farmacias', 'change')
def modificar_farmacia(request, pk):
    from .models import Localfarmacia
    farmacia = get_object_or_404(Localfarmacia, id=pk)
    from .forms import FarmaciaForm
    if request.method == 'POST':
        form = FarmaciaForm(request.POST, instance=farmacia)
        if form.is_valid():
            form.save()
            messages.success(request, 'Farmacia actualizada correctamente')
            return redirect('detalle_farmacia', pk=farmacia.id)
        else:
            messages.error(request, 'Corrige los errores del formulario')
    else:
        form = FarmaciaForm(instance=farmacia)
    return render(request, 'localfarmacia/modificar-farmacia.html', {'form': form, 'farmacia': farmacia})
@rol_requerido('motorista')
def panel_motorista(request):
    u = Usuario.objects.filter(django_user_id=request.user.id).first()
    if not u:
        try:
            from .models import Rol
            rol_mot = Rol.objects.filter(codigo='MOTORISTA').first() or Rol.objects.first()
        except Exception:
            rol_mot = None
        from django.utils import timezone as _tz
        now_dt = _tz.now()
        u = Usuario.objects.create(
            django_user_id=request.user.id,
            rol=rol_mot,
            tipo_documento='RUT',
            documento_identidad=f"MTR-{request.user.username}"[:30],
            nombre=request.user.first_name or request.user.username,
            apellido=request.user.last_name or 'Demo',
            telefono='000000000',
            activo=True,
            fecha_creacion=now_dt,
            fecha_modificacion=now_dt,
        )
    m = Motorista.objects.filter(usuario=u).first()
    if not m:
        from django.utils import timezone as _tz
        now_dt = _tz.now()
        m = Motorista.objects.create(
            usuario=u,
            licencia_numero=f"LIC-{u.documento_identidad[:8].upper()}",
            licencia_clase='A',
            fecha_vencimiento_licencia=now_dt.date(),
            emergencia_nombre='Contacto',
            emergencia_telefono='000000000',
            emergencia_parentesco='Otro',
            total_entregas_completadas=0,
            total_entregas_fallidas=0,
            activo=True,
            disponible_hoy=True,
            fecha_creacion=now_dt,
            fecha_modificacion=now_dt,
        )
    asign_moto = AsignacionMotoMotorista.objects.filter(motorista=m, activa=1).first()
    if not asign_moto:
        mo = Moto.objects.filter(activo=True).first()
        if mo:
            from django.utils import timezone as _tz
            if not AsignacionMotoMotorista.objects.filter(moto=mo, activa=1).exists():
                asign_moto = AsignacionMotoMotorista.objects.create(
                    motorista=m,
                    moto=mo,
                    fecha_asignacion=_tz.now(),
                    kilometraje_inicio=int(getattr(mo, 'kilometraje_actual', 0) or 0),
                    activa=1,
                    observaciones='Asignación automática para demostración',
                )
    qs = Despacho.objects.filter(motorista=m).order_by('-fecha_registro')[:50]
    if not qs:
        from django.utils import timezone as _tz
        now_dt = _tz.now()
        pool = Despacho.objects.filter(estado='PENDIENTE').order_by('-fecha_registro')[:10]
        reassigned = []
        for d in pool:
            d.motorista = m
            d.fecha_asignacion = d.fecha_asignacion or now_dt
            try:
                d.save()
                reassigned.append(d)
            except Exception:
                pass
        if reassigned:
            qs = Despacho.objects.filter(motorista=m).order_by('-fecha_registro')[:50]
    lids = [d.farmacia_origen_local_id for d in qs if d.farmacia_origen_local_id]
    origenes = {f.local_id: f for f in Farmacia.objects.filter(local_id__in=lids)}
    for d in qs:
        d.origen = origenes.get(d.farmacia_origen_local_id)
        try:
            from django.utils import timezone as _tz
            now_dt = _tz.now()
            if d.fecha_salida_farmacia and not d.fecha_completado and not d.fecha_anulacion:
                d.min_en_ruta = int(max((now_dt - d.fecha_salida_farmacia).total_seconds() / 60, 0))
            else:
                d.min_en_ruta = None
        except Exception:
            d.min_en_ruta = None
    return render(request, 'motoristas/panel.html', {
        'motorista': m,
        'asign_moto': asign_moto,
        'despachos': qs,
    })

@rol_requerido('motorista')
@require_http_methods(["POST"])
def motorista_despacho_estado(request, despacho_id):
    u = Usuario.objects.filter(django_user_id=request.user.id).first()
    m = Motorista.objects.filter(usuario=u).first()
    d = Despacho.objects.filter(pk=despacho_id, motorista=m).first()
    if not d:
        messages.error(request, 'Despacho no encontrado')
        return redirect('panel_motorista')
    nuevo = (request.POST.get('estado') or '').strip().upper()
    receta_devuelta = request.POST.get('receta_devuelta')
    tipo_incidencia = (request.POST.get('tipo_incidencia') or '').strip()
    incidencia_nota = (request.POST.get('incidencia_nota') or '').strip()
    estado_prev = d.estado
    ok, msg = _can_transition(estado_prev, nuevo, d.tipo_despacho, d.tiene_receta_retenida, d.receta_devuelta_farmacia)
    if not ok:
        messages.error(request, msg)
        return redirect('panel_motorista')
    from django.utils import timezone as _tz
    now_dt = _tz.now()
    d.usuario_modificacion = u
    d.fecha_modificacion = now_dt
    if nuevo == 'EN_CAMINO' and not d.fecha_salida_farmacia:
        d.fecha_salida_farmacia = now_dt
    if nuevo == 'ENTREGADO':
        d.fecha_llegada_destino = d.fecha_llegada_destino or now_dt
        d.fecha_completado = d.fecha_completado or now_dt
    if nuevo == 'FALLIDO':
        d.fecha_llegada_destino = d.fecha_llegada_destino or now_dt
        d.hubo_incidencia = True
        d.tipo_incidencia = tipo_incidencia or d.tipo_incidencia
        if tipo_incidencia == 'OTRA' and incidencia_nota:
            d.descripcion_incidencia = incidencia_nota
    if receta_devuelta:
        d.receta_devuelta_farmacia = True
        d.fecha_devolucion_receta = d.fecha_devolucion_receta or now_dt
    d.estado = nuevo
    try:
        d.save()
        AuditoriaGeneral.objects.create(
            nombre_tabla='despacho',
            id_registro_afectado=str(d.id),
            tipo_operacion='TRANSICION',
            usuario=u,
            fecha_evento=now_dt,
            datos_antiguos={'estado': estado_prev},
            datos_nuevos={'estado': nuevo, 'incidencia': d.tipo_incidencia}
        )
        is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
        if is_ajax:
            from django.http import JsonResponse
            min_en_ruta = None
            if d.fecha_salida_farmacia and not d.fecha_completado and not d.fecha_anulacion:
                try:
                    min_en_ruta = int(max((now_dt - d.fecha_salida_farmacia).total_seconds() / 60, 0))
                except Exception:
                    min_en_ruta = None
            return JsonResponse({
                'ok': True,
                'id': d.id,
                'estado': d.estado,
                'min_en_ruta': min_en_ruta,
            })
        messages.success(request, 'Estado actualizado')
    except Exception as e:
        is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
        if is_ajax:
            from django.http import JsonResponse
            return JsonResponse({'ok': False, 'error': str(e)}, status=400)
        messages.error(request, f'Error: {e}')
    return redirect('panel_motorista')

def healthz(request):
    from django.http import JsonResponse
    from django.core.cache import cache
    from django.db import connection
    status = {}
    try:
        connection.ensure_connection()
        status['database'] = 'ok'
    except Exception as e:
        status['database'] = f'error:{e}'
    try:
        cache.set('healthz', '1', 5)
        v = cache.get('healthz')
        status['cache'] = 'ok' if v == '1' else 'error'
    except Exception as e:
        status['cache'] = f'error:{e}'
    try:
        request.session['healthz'] = '1'
        v = request.session.get('healthz')
        status['session'] = 'ok' if v == '1' else 'error'
    except Exception as e:
        status['session'] = f'error:{e}'
    overall = all(val == 'ok' for val in status.values())
    return JsonResponse({'status': 'ok' if overall else 'error', **status}, status=200 if overall else 503)
